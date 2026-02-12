#!/usr/bin/env python3
"""
Unified program to check for invalid DUPLICATA and SEUNO records.
Queries APR_CAPA, APR_TITULOS, and CEDENTE tables with STATUS column.

This program combines:
1. Invalid DUPLICATA validation (from test_query.py / duplicatas_invalidas)
2. Invalid SEUNO validation (from query_pre_impresso_with_propostas.py / seuno_invalidos)
3. STATUS column from APR_CAPA table
4. Status filter: Only processes records with status >= "Aguardando Analista"

Features:
- Validates DUPLICATA format (NFE + separator + sequential number)
- Validates SEUNO with verification digit algorithm
- Filters by status progression (Aguardando Analista or further)
- Exports invalid records to Excel (separate files for DUPLICATA and SEUNO)
- Sends Slack notifications with Excel attachments
- Stores valid records in MariaDB for tracking
- Production mode by default (use --dry-run to preview without sending alerts)
"""

import json
import sys
import re
import hashlib
from pathlib import Path
from datetime import datetime, date, timedelta
from sqlalchemy import create_engine, or_, and_, func, cast, Date
from sqlalchemy.orm import sessionmaker
from models import Cedente, APRTitulos, APRCapa, Produto, ProdutoCedente
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import requests
import pymysql

# Import monitoring helpers
from monitoring_helpers import (
    log_invalid_records,
    auto_resolve_invalid_records,
    track_all_status_changes_from_source,
    track_proposal_products,
    log_program_execution,
    update_daily_summary,
    cleanup_old_monitoring_data
)
from models_mariadb import MariaDBBase


def get_most_recent_weekday(target_date=None):
    """
    Get the most recent weekday (Monday-Friday).
    If target_date is a weekend, return the most recent Friday.
    """
    if target_date is None:
        target_date = date.today()

    # weekday(): Monday=0, Sunday=6
    # Saturday=5, Sunday=6
    while target_date.weekday() >= 5:  # Saturday or Sunday
        target_date -= timedelta(days=1)

    return target_date


def load_config() -> dict:
    """Load database configuration from databases_config.json."""
    config_paths = [
        Path.cwd() / "databases_config.json",  # Current working directory (for standalone executables)
        Path(__file__).parent / "databases_config.json",
        Path(__file__).parent.parent / "databases_config.json",
    ]

    for config_path in config_paths:
        if config_path.exists():
            with open(config_path, 'r') as f:
                return json.load(f)

    raise FileNotFoundError(
        f"Configuration file not found in: {[str(p) for p in config_paths]}"
    )


def load_slack_config() -> dict:
    """Load Slack configuration from slack_config.json."""
    config_paths = [
        Path.cwd() / "slack_config.json",  # Current working directory (for standalone executables)
        Path(__file__).parent / "slack_config.json",
        Path(__file__).parent.parent / "slack_config.json",
    ]

    for config_path in config_paths:
        if config_path.exists():
            with open(config_path, 'r') as f:
                return json.load(f)

    # Return default config if file not found
    return {
        "slack": {
            "webhook_url": None,
            "channel": "#alerts",
            "username": "BMA Alert Bot"
        }
    }


def create_connection_string(config: dict) -> str:
    """Create SQLAlchemy connection string for MSSQL."""
    cfg = config['databases']['mssql']
    server = cfg.get('server', 'localhost')
    port = cfg.get('port', 1433)
    database = cfg.get('scheme', 'BMA')
    username = cfg.get('user', '')
    password = cfg.get('password', '')

    # Using pymssql driver
    return f"mssql+pymssql://{username}:{password}@{server}:{port}/{database}"


def create_mariadb_connection(config):
    """
    Create MariaDB connection for storing valid records.

    Args:
        config: Database configuration dictionary

    Returns:
        pymysql connection object
    """
    db_config = config['databases']['mariadb']

    connection = pymysql.connect(
        host=db_config['server'],
        port=db_config['port'],
        user=db_config['user'],
        password=db_config['password'],
        database=db_config['scheme'],
        charset='utf8mb4',
        cursorclass=pymysql.cursors.DictCursor
    )

    return connection


def create_mariadb_session(config):
    """
    Create SQLAlchemy session for MariaDB (for monitoring functions).

    Args:
        config: Database configuration dictionary

    Returns:
        SQLAlchemy session object
    """
    db_config = config['databases']['mariadb']

    connection_string = (
        f"mysql+pymysql://{db_config['user']}:{db_config['password']}"
        f"@{db_config['server']}:{db_config['port']}/{db_config['scheme']}"
        f"?charset=utf8mb4"
    )

    engine = create_engine(connection_string, echo=False)
    Session = sessionmaker(bind=engine)
    return Session()


def store_valid_records_to_mariadb(valid_records, mariadb_conn, dry_run=True):
    """
    Store or update valid records in MariaDB table.
    Uses composite key: (DATA, PROPOSTA, CEDENTE, RAMO)

    IMPORTANT: The following fields are NEVER updated (only set on insert):
    - DATA, PROPOSTA, CEDENTE, RAMO (composite key)
    - GERENTE, EMPRESA (fixed fields)

    Only these fields can be updated:
    - STATUS, QTD_APROVADOS, VLR_APROVADOS, VALOR_TITULOS, QTD_TITULOS, update_count

    Args:
        valid_records: List of valid record dictionaries
        mariadb_conn: MariaDB connection object
        dry_run: If True, don't actually insert/update

    Returns:
        Tuple of (inserted_count, updated_count)
    """
    if not valid_records:
        return 0, 0

    # Group records by composite key to aggregate title values
    aggregated = {}

    for record in valid_records:
        key = (
            record['DATA'],
            record['PROPOSTA'],
            record['CEDENTE'],
            record['RAMO']
        )

        if key not in aggregated:
            aggregated[key] = {
                'DATA': record['DATA'],
                'PROPOSTA': record['PROPOSTA'],
                'CEDENTE': record['CEDENTE'],
                'RAMO': record['RAMO'],
                'GERENTE': record['GERENTE'],
                'EMPRESA': record['EMPRESA'],
                'STATUS': record['STATUS'],
                'QTD_APROVADOS': record['QTD_APROVADOS'],
                'VLR_APROVADOS': record['VLR_APROVADOS'],
                'VALOR_TITULOS': 0.0,
                'QTD_TITULOS': 0
            }

        # Aggregate title values
        aggregated[key]['VALOR_TITULOS'] += record['VALOR']
        aggregated[key]['QTD_TITULOS'] += 1

    insert_count = 0
    update_count = 0

    # IMPORTANT: ON DUPLICATE KEY UPDATE does NOT update:
    # - DATA, PROPOSTA, CEDENTE, RAMO (composite key - cannot be updated)
    # - GERENTE, EMPRESA (fixed fields - should not change)
    # - is_processado (processing status - should not be reset on update)
    # Only updates: STATUS, QTD_APROVADOS, VLR_APROVADOS, VALOR_TITULOS, QTD_TITULOS, update_count
    upsert_sql = """
    INSERT INTO apr_valid_records
        (DATA, PROPOSTA, CEDENTE, RAMO, GERENTE, EMPRESA, STATUS,
         QTD_APROVADOS, VLR_APROVADOS, VALOR_TITULOS, QTD_TITULOS, is_processado, update_count)
    VALUES
        (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, 0, 0)
    ON DUPLICATE KEY UPDATE
        STATUS = VALUES(STATUS),
        QTD_APROVADOS = VALUES(QTD_APROVADOS),
        VLR_APROVADOS = VALUES(VLR_APROVADOS),
        VALOR_TITULOS = VALUES(VALOR_TITULOS),
        QTD_TITULOS = VALUES(QTD_TITULOS),
        update_count = update_count + 1
    """

    if dry_run:
        print(f"\n[DRY RUN] Would insert/update {len(aggregated)} records in MariaDB")
        print(f"  Sample records (first 5):")
        for i, (key, rec) in enumerate(list(aggregated.items())[:5], 1):
            print(f"    {i}. DATA={rec['DATA']}, PROP={rec['PROPOSTA']}, "
                  f"CEDENTE={rec['CEDENTE'][:20]}, STATUS={rec['STATUS']}, "
                  f"QTD_TIT={rec['QTD_TITULOS']}, VLR={rec['VALOR_TITULOS']:.2f}")
        return 0, 0

    with mariadb_conn.cursor() as cursor:
        for key, rec in aggregated.items():
            cursor.execute(upsert_sql, (
                rec['DATA'],
                rec['PROPOSTA'],
                rec['CEDENTE'],
                rec['RAMO'],
                rec['GERENTE'],
                rec['EMPRESA'],
                rec['STATUS'],
                rec['QTD_APROVADOS'],
                rec['VLR_APROVADOS'],
                rec['VALOR_TITULOS'],
                rec['QTD_TITULOS']
            ))

            # Check if it was an insert or update
            if cursor.rowcount == 1:
                insert_count += 1
            elif cursor.rowcount == 2:  # ON DUPLICATE KEY UPDATE affects 2 rows
                update_count += 1

        mariadb_conn.commit()

    return insert_count, update_count


# ============================================================================
# NFECHAVE VALIDATION FUNCTIONS
# ============================================================================

def validate_nfechave(nfechave: str) -> bool:
    """
    Validate if NFEChave field is not empty, null, or whitespace-only.

    Args:
        nfechave: The NFEChave field value

    Returns:
        True if valid (not empty), False otherwise
    """
    if not nfechave:
        return False

    # Check if it's only whitespace
    if not nfechave.strip():
        return False

    return True


def get_invalid_nfechave_reason(produto: str) -> str:
    """
    Get the reason why NFEChave is invalid (in Brazilian Portuguese).

    Args:
        produto: The PRODUTO name

    Returns:
        String describing the error reason in Portuguese
    """
    return f"{produto} sem chave de NF"


# ============================================================================
# DUPLICATA VALIDATION FUNCTIONS
# ============================================================================

def validate_duplicata_format(duplicata: str, nfe: int) -> bool:
    """
    Validate DUPLICATA nomenclature.

    Format should be: NFE (with or without leading zeros) + separator (/ or -) + sequential number
    Examples:
        - 021467-1, 021467-2 (valid for NFE=21467)
        - 21467-1, 21467-2 (valid for NFE=21467)
        - 140917/01, 140917/02 (valid for NFE=140917)

    Args:
        duplicata: The DUPLICATA field value
        nfe: The NFE integer value

    Returns:
        True if valid format, False otherwise
    """
    if not duplicata or not nfe:
        return False

    # Pattern: NFE (with optional leading zeros) + separator (/ or -) + sequential number (1+ digits)
    # The NFE part should match the actual NFE value (with or without leading zeros)
    pattern = rf'^0*{nfe}[-/]\d+$'

    return bool(re.match(pattern, duplicata.strip()))


def get_invalid_duplicata_reason(duplicata: str, nfe: int) -> str:
    """
    Get the reason why DUPLICATA format is invalid (in Brazilian Portuguese).

    Args:
        duplicata: The DUPLICATA field value
        nfe: The NFE integer value

    Returns:
        String describing the error reason in Portuguese
    """
    if not duplicata:
        return "DUPLICATA vazio ou nulo"

    if not nfe:
        return "NFE vazio ou nulo"

    duplicata = duplicata.strip()

    # Check if it contains the NFE number
    if str(nfe) not in duplicata and not re.search(rf'0*{nfe}', duplicata):
        return f"DUPLICATA não contém o número da NFE ({nfe})"

    # Check if it has a separator
    if '-' not in duplicata and '/' not in duplicata:
        return "Falta separador (- ou /) entre NFE e número sequencial"

    # Check if NFE is at the beginning
    if not re.match(rf'^0*{nfe}', duplicata):
        return f"NFE ({nfe}) não está no início da DUPLICATA"

    # Check if there's a sequential number after separator
    if not re.search(r'[-/]\d+$', duplicata):
        return "Falta número sequencial após o separador"

    # Generic error
    return "Formato inválido - esperado: NFE + separador (- ou /) + número sequencial"


def validate_duplicata_format_simple(duplicata: str) -> bool:
    """
    Validate DUPLICATA format without NFE validation.
    Used for specific product types that don't require NFE.

    Format: any text + separator (/ or -) + sequential number

    Args:
        duplicata: The DUPLICATA field value

    Returns:
        True if has separator and sequential number, False otherwise
    """
    if not duplicata:
        return False

    # Pattern: anything + separator (/ or -) + sequential number (1+ digits)
    pattern = r'.+[-/]\d+$'
    return bool(re.match(pattern, duplicata.strip()))


def get_invalid_duplicata_reason_simple(duplicata: str) -> str:
    """
    Get reason why DUPLICATA format is invalid (simple validation, no NFE).

    Args:
        duplicata: The DUPLICATA field value

    Returns:
        String describing the error reason in Portuguese
    """
    if not duplicata:
        return "DUPLICATA vazio ou nulo"

    duplicata = duplicata.strip()

    # Check if it has a separator
    if '-' not in duplicata and '/' not in duplicata:
        return "Falta separador (- ou /) e número sequencial"

    # Check if there's a sequential number after separator
    if not re.search(r'[-/]\d+$', duplicata):
        return "Falta número sequencial após o separador"

    return "Formato inválido - esperado: separador (- ou /) + número sequencial"


def validate_cheque_proposals(session, validated_results, target_date):
    """
    Validate proposals that contain ONLY CHEQUE products.

    Rules:
    - If a proposal has ONLY CHEQUE products
    - Count total number of titles (QTD_TITULOS)
    - If > 100 titles: INVALID (send alert)
    - If <= 100 titles: VALID

    Args:
        session: SQLAlchemy session for MSSQL
        validated_results: List of validated records
        target_date: Date being processed

    Returns:
        Tuple of (invalid_cheque_records, updated_validated_results)
    """
    from sqlalchemy import func
    from collections import defaultdict

    print("\n" + "=" * 140)
    print("VALIDATING CHEQUE-ONLY PROPOSALS")
    print("=" * 140)

    # Group records by proposal to check if they have ONLY CHEQUE
    proposal_products = defaultdict(set)
    proposal_info = {}  # Store proposal metadata

    for record in validated_results:
        key = (record['DATA'], record['PROPOSTA'], record['CEDENTE'])
        produto = (record.get('PRODUTO') or '').upper()
        proposal_products[key].add(produto)

        if key not in proposal_info:
            proposal_info[key] = {
                'GERENTE': record.get('GERENTE'),
                'DATA': record['DATA'],
                'PROPOSTA': record['PROPOSTA'],
                'CEDENTE': record['CEDENTE'],
                'EMPRESA': record.get('EMPRESA'),
                'STATUS': record.get('STATUS'),
                'RAMO': record.get('RAMO'),
            }

    # Find proposals with ONLY CHEQUE
    cheque_only_proposals = []
    for key, products in proposal_products.items():
        if products == {'CHEQUE'}:
            cheque_only_proposals.append(key)

    print(f"Found {len(cheque_only_proposals)} proposals with ONLY CHEQUE products")

    if not cheque_only_proposals:
        print("✓ No CHEQUE-only proposals to validate")
        return [], validated_results

    # Query to count titles for each CHEQUE-only proposal
    invalid_cheque_records = []

    for data, proposta, cedente in cheque_only_proposals:
        # Count titles for this proposal
        count_query = session.query(
            func.count(APRTitulos.TITULO).label('QTD_TITULOS')
        ).join(
            APRCapa,
            and_(
                APRTitulos.DATA == APRCapa.DATA,
                APRTitulos.NUMERO == APRCapa.NUMERO
            )
        ).filter(
            cast(APRCapa.DATA, Date) == data,
            APRCapa.NUMERO == proposta,
            APRCapa.CEDENTE == cedente
        )

        result = count_query.first()
        qtd_titulos = result.QTD_TITULOS if result else 0

        # Check if exceeds limit
        if qtd_titulos > 100:
            info = proposal_info[(data, proposta, cedente)]
            invalid_cheque_records.append({
                'GERENTE': info['GERENTE'],
                'PROPOSTA': proposta,
                'DATA': data,
                'CEDENTE': cedente,
                'EMPRESA': info['EMPRESA'],
                'STATUS': info['STATUS'],
                'RAMO': info['RAMO'],
                'PRODUTO': 'CHEQUE',
                'QTD_TITULOS': qtd_titulos,
                'MOTIVO_INVALIDO': f'Proposta CHEQUE com {qtd_titulos} títulos (limite: 100)',
                'NFECHAVE_VALID': True,  # CHEQUE doesn't need NFEChave
                'DUPLICATA_VALID': True,  # CHEQUE doesn't need DUPLICATA
                'SEUNO_VALID': True,      # CHEQUE doesn't need SEUNO
                'IS_VALID': False,        # But it's invalid due to quantity
                'SEUNO': None,
                'DUPLICATA': None,
                'NFE': None,
            })
            print(f"  ⚠️  INVALID: {cedente} - Proposta {proposta} - {qtd_titulos} títulos (> 100)")
        else:
            print(f"  ✓ VALID: {cedente} - Proposta {proposta} - {qtd_titulos} títulos (<= 100)")

    # Mark CHEQUE-only proposals as valid in validated_results if they pass the check
    # (they were already validated, we just need to mark the invalid ones)
    updated_results = []
    for record in validated_results:
        key = (record['DATA'], record['PROPOSTA'], record['CEDENTE'])

        # If this is a CHEQUE-only proposal that's invalid, mark it
        if key in cheque_only_proposals:
            # Check if it's in the invalid list
            is_invalid_cheque = any(
                inv['DATA'] == record['DATA'] and
                inv['PROPOSTA'] == record['PROPOSTA'] and
                inv['CEDENTE'] == record['CEDENTE']
                for inv in invalid_cheque_records
            )

            if is_invalid_cheque:
                # Skip this record from validated_results (it's now invalid)
                continue

        updated_results.append(record)

    print(f"\n✓ CHEQUE validation complete: {len(invalid_cheque_records)} invalid proposals found")
    print("=" * 140)

    return invalid_cheque_records, updated_results


# ============================================================================
# SEUNO VALIDATION FUNCTIONS
# ============================================================================

def calculate_verification_digit(seuno):
    """
    Calculate the verification digit for a SEUNO.

    Algorithm:
    1. Form seuno_vector: [1, 9] + first 11 digits of SEUNO
    2. factor_vector: [2, 7, 6, 5, 4, 3, 2, 7, 6, 5, 4, 3, 2]
    3. dot_product = sum(seuno_vector[i] * factor_vector[i])
    4. quotient = dot_product // 11
    5. pre_last_remainder = dot_product - (quotient * 11)  [this is the modulo/remainder]
    6. Calculate verification_digit:
       - if (11 - pre_last_remainder) == 11: verification_digit = 0
       - if (11 - pre_last_remainder) == 10: verification_digit = 'P'
       - else: verification_digit = 11 - pre_last_remainder

    Args:
        seuno: string representing the SEUNO (e.g., '518300121527')

    Returns:
        verification_digit as string (digit or 'P')
    """
    if not seuno or len(seuno) < 11:
        return ''

    # Extract first 11 digits from SEUNO (ignore non-digit characters)
    seuno_digits = [int(c) for c in seuno if c.isdigit()][:11]

    if len(seuno_digits) < 11:
        return ''

    # Form seuno_vector: [1, 9] + first 11 digits
    seuno_vector = [1, 9] + seuno_digits

    # Fixed factor_vector
    factor_vector = [2, 7, 6, 5, 4, 3, 2, 7, 6, 5, 4, 3, 2]

    # Calculate dot product
    dot_product = sum(s * f for s, f in zip(seuno_vector, factor_vector))

    # Calculate quotient (integer division)
    quotient = dot_product // 11

    # Calculate pre_last_remainder (this is the modulo: dot_product % 11)
    pre_last_remainder = dot_product - (quotient * 11)

    # Calculate verification_digit based on rules
    result = 11 - pre_last_remainder

    if result == 11:
        verification_digit = '0'
    elif result == 10:
        verification_digit = 'P'
    else:
        verification_digit = str(result)

    return verification_digit


def validate_seuno(seuno, range_value, calculated_verif):
    """
    Validate SEUNO according to business rules.

    Rules:
    1. SEUNO must start with the RANGE value
    2. After the first 11 digits, it must contain a valid verification digit (0, P, or 2-9)

    Args:
        seuno: string representing the SEUNO (e.g., '518300121527')
        range_value: string representing the expected range (e.g., '5183')
        calculated_verif: string representing the calculated verification digit

    Returns:
        tuple (is_valid: bool, motivo_invalido: str)
    """
    if not seuno:
        return False, "SEUNO vazio ou nulo"

    # Remove non-digit characters for validation (except P)
    seuno_clean = ''.join(c for c in seuno if c.isdigit() or c.upper() == 'P')

    if len(seuno_clean) < 12:
        return False, "SEUNO com menos de 12 caracteres"

    # Rule 1: SEUNO must start with RANGE value
    if not seuno_clean.startswith(range_value):
        return False, f"SEUNO não inicia com a faixa {range_value}"

    # Rule 2: Position 12 (index 11) must contain a valid verification digit
    actual_verif = seuno_clean[11] if len(seuno_clean) > 11 else ''

    # Valid verification digits: 0, P, 1-9
    valid_digits = {'0', 'P', '1', '2', '3', '4', '5', '6', '7', '8', '9'}

    if actual_verif not in valid_digits:
        return False, f"Dígito verificador inválido na posição 12: '{actual_verif}'"

    # Check if the actual verification digit matches the calculated one
    if actual_verif != calculated_verif:
        return False, f"Dígito verificador incorreto: esperado '{calculated_verif}', encontrado '{actual_verif}'"

    return True, ""


def parse_pre_impresso_ranges(obs1_text: str) -> list:
    """
    Parse pre-impresso ranges from obs1 field.

    Returns list of (company, range) tuples.

    STRICT RULES:
    - Must contain "pré" or "pre" followed by "impresso"
    - Must contain company name: BMA, BMA FIDC, BMA INTER, or BMA SEC (normalized to BMA)
    - Must contain a range number (digits) after the company name
    - MUST IGNORE if contains "parametrização" or variants

    SPECIAL CASES:
    - If contains "pré/pre" + "impresso" + "webfact" => range 200
    - If contains "parametrização" => IGNORE (return empty list)
    - "BMA SEC" => normalized to "BMA"
    - "Fidc" (without BMA) => normalized to "BMA FIDC"
    - "Matriz faixa X" and "Filial Y" => capture both X and Y

    Supported formats:
    1. "Boleto Pré impresso BMA FIDC - Faixa N/N - 02014"
    2. "Pré-impresso BMA INTER - Faixa 5101"
    3. "Boleto pré-impresso - BMA FIDC faixa n/n 5233"
    4. "Pré impresso BMA FIDC - Faixa - 0850"
    5. "Boleto Pré Impresso - Via Webfact" => range 200
    6. "Boleto Pré-Impresso BMA Inter: faixa 6003" (colon after company)
    7. Multi-line format (concatenated with spaces)
    8. "Pré-impresso BMA SEC - Faixa 5125" => BMA:5125
    9. "Pré-impresso Fidc Faixa 5325" => BMA FIDC:5325
    10. "Matriz faixa 3 - Filial 5061" => BMA:3, BMA:5061
    """
    if not obs1_text:
        return []

    ranges = []

    # Handle multi-line format: concatenate lines with spaces
    obs1_normalized = ' '.join(obs1_text.split('\n'))
    obs1_upper = obs1_normalized.upper()

    # Check if text contains "pré-impresso" or "pre-impresso"
    has_pre_impresso = ('PRE' in obs1_upper or 'PRÉ' in obs1_upper) and 'IMPRESSO' in obs1_upper

    if not has_pre_impresso:
        return []

    # SPECIAL CASE 1: IGNORE if "parametrização" appears near "pré-impresso"
    # Only ignore if parametrização is in the same context as pré-impresso
    # Pattern: "pré-impresso" followed by "parametrização" within ~50 characters
    if re.search(r'PR[EÉ][\s\-]*[IÌ]MPRESSO.{0,50}PARAMETRI[ZS]A[ÇC][ÃA]O', obs1_upper):
        return []  # Ignore cases being configured
    # Also check reverse: "parametrização" followed by "pré-impresso" within ~50 characters
    if re.search(r'PARAMETRI[ZS]A[ÇC][ÃA]O.{0,50}PR[EÉ][\s\-]*[IÌ]MPRESSO', obs1_upper):
        return []  # Ignore cases being configured

    # SPECIAL CASE 2: Check for webfact with pre-impresso
    # These cases get assigned range 200
    has_webfact = 'WEBFACT' in obs1_upper

    if has_webfact:
        ranges.append(("BMA", "200"))
        return ranges  # Return immediately for webfact cases

    # First, try to find the company name mentioned in the text (for Matriz/Filial inheritance)
    # This is used even if the company doesn't have a direct range number
    first_company = None
    company_mention_pattern = r'(BMA\s+(?:FIDC|INTER|SEC)|FIDC|BMA)'
    company_mention_match = re.search(company_mention_pattern, obs1_upper)

    if company_mention_match:
        company = company_mention_match.group(1).strip()
        # Normalize company name
        if 'SEC' in company:
            first_company = 'BMA'  # BMA SEC => BMA
        elif 'FIDC' in company:
            first_company = 'BMA FIDC'
        elif 'INTER' in company:
            first_company = 'BMA INTER'
        else:
            first_company = 'BMA'

    # Main pattern: Find company + range combinations
    # Supports: BMA FIDC, BMA INTER, BMA SEC, BMA, FIDC (standalone)
    company_range_pattern = r'(BMA\s+(?:FIDC|INTER|SEC)|FIDC|BMA)[\s\-:\.]*(?:FAIXA)?[\s\-:]*(?:[N/]+)?[\s\-:]*(\d{1,5})'

    for match in re.finditer(company_range_pattern, obs1_upper):
        company = match.group(1).strip()
        range_val = match.group(2).strip()

        if range_val and company:  # Both company and range must be present
            # Normalize company name
            if 'SEC' in company:
                company_normalized = 'BMA'  # BMA SEC => BMA
            elif 'FIDC' in company:
                company_normalized = 'BMA FIDC'
            elif 'INTER' in company:
                company_normalized = 'BMA INTER'
            else:
                company_normalized = 'BMA'

            # Avoid duplicates
            if (company_normalized, range_val) not in ranges:
                ranges.append((company_normalized, range_val))

    # SPECIAL CASE 3: Handle "Matriz faixa X" and "Filial Y" patterns
    # These inherit the first company found, or default to BMA
    # Example: "Pré-impresso BMA FIDC - Matriz faixa 3 - Filial 5061"
    #          => BMA FIDC:3, BMA FIDC:5061
    matriz_filial_pattern = r'(?:MATRIZ|FILIAL)[\s\-]*(?:FAIXA)?[\s\-]*(\d{1,5})'

    for match in re.finditer(matriz_filial_pattern, obs1_upper):
        range_val = match.group(1).strip()

        if range_val:
            # Use first company found, or default to BMA
            company_for_matriz_filial = first_company if first_company else 'BMA'

            # Avoid duplicates
            if (company_for_matriz_filial, range_val) not in ranges:
                ranges.append((company_for_matriz_filial, range_val))

    return ranges


# ============================================================================
# STATUS FILTER FUNCTIONS
# ============================================================================

# Define the status progression order
# Status flows from left to right (earlier to later stages)
STATUS_PROGRESSION = [
    "Aguardando Analista",
    "Aguardando análise",
    "Em análise",
    "Aguardando Aprovação",
    "Aprovado",
    "Enviado para Assinar",
    "Assinado",
    "Liberado",
    "Finalizado",
    "Pago"  # Proposals that have been paid
]


def is_status_equal_or_further(status: str, reference_status: str = "Aguardando Analista") -> bool:
    """
    Check if a status is equal to or further than a reference status in the workflow.

    Args:
        status: The status to check
        reference_status: The reference status to compare against (default: "Aguardando Analista")

    Returns:
        True if status is equal or further, False otherwise
    """
    if not status:
        return False

    # Normalize status strings (case-insensitive, strip whitespace)
    status_normalized = status.strip()
    reference_normalized = reference_status.strip()

    # Try to find the status in the progression list (case-insensitive)
    status_index = None
    reference_index = None

    for i, prog_status in enumerate(STATUS_PROGRESSION):
        if prog_status.lower() == status_normalized.lower():
            status_index = i
        if prog_status.lower() == reference_normalized.lower():
            reference_index = i

    # If we found both statuses in the progression, compare their positions
    if status_index is not None and reference_index is not None:
        return status_index >= reference_index

    # If status not found in progression but reference is "Aguardando Analista",
    # check for partial match (for variations like "Aguardando análise")
    if reference_normalized.lower() == "aguardando analista":
        # Accept any status that contains "aguardando" or is further in the workflow
        if "aguardando" in status_normalized.lower():
            return True
        # If status is not in progression and doesn't contain "aguardando", assume it's earlier
        if status_index is None:
            return False

    # Default: if we can't determine, return False (conservative approach)
    return False


# ============================================================================
# MAIN QUERY FUNCTION
# ============================================================================

def query_apr_invalidos_with_status(session, target_date=None, apply_status_filter=True):
    """
    Query APR_CAPA, APR_TITULOS, and CEDENTE tables.
    Returns records with both DUPLICATA and SEUNO validation, plus STATUS.

    Args:
        session: SQLAlchemy session
        target_date: Date to query (default: today)
        apply_status_filter: If True, only include records with status >= "Aguardando Analista"

    Returns:
        List of dictionaries with validation results
    """
    if target_date is None:
        target_date = date.today()

    # Ensure we use a weekday
    original_date = target_date
    target_date = get_most_recent_weekday(target_date)

    if original_date != target_date:
        print(f"Note: {original_date} is a weekend. Using most recent weekday: {target_date}")

    print(f"\nQuerying APR data for date: {target_date}")
    if apply_status_filter:
        print("Status Filter: Only records with status >= 'Aguardando Analista'")
    print("=" * 140)

    # Step 1: Get cedentes with pre-impresso and their ranges
    print("\nStep 1: Extracting cedentes with pre-impresso...")
    cedentes_with_pre_impresso = session.query(Cedente).filter(
        Cedente.obs1 != None,
        or_(
            Cedente.obs1.like('%pre%impresso%'),
            Cedente.obs1.like('%pré%impresso%'),
            Cedente.obs1.like('%pre%ìmpresso%'),
            Cedente.obs1.like('%pré%ìmpresso%'),
        )
    ).all()

    # Build a mapping: cedente apelido -> list of (company, range) pairs
    cedente_ranges = {}
    for cedente in cedentes_with_pre_impresso:
        ranges = parse_pre_impresso_ranges(cedente.obs1)
        if ranges:
            cedente_ranges[cedente.apelido] = ranges

    print(f"Found {len(cedente_ranges)} cedentes with pre-impresso ranges")

    # Step 2: Query APR_CAPA joined with APR_TITULOS and CEDENTE
    # Include STATUS from APR_CAPA and RAMO from CEDENTE
    # Also join with Product tables to get PRODUTO name
    print("\nStep 2: Querying APR_CAPA + APR_TITULOS + CEDENTE...")

    # Query 2a: Records WITH NFEChave (existing logic)
    query_with_nfe = session.query(
        APRCapa.GERENTE,
        APRCapa.NUMERO.label('PROPOSTA'),
        APRCapa.DATA,
        APRCapa.CEDENTE,
        APRCapa.empresa,
        APRCapa.status_atual.label('STATUS'),
        Cedente.ramo.label('RAMO'),
        APRTitulos.SEUNO,
        APRTitulos.TITULO.label('DUPLICATA'),
        APRTitulos.NFEChave,
        Produto.Descritivo.label('PRODUTO'),
        Cedente.apelido
    ).join(
        APRTitulos,
        and_(
            APRCapa.DATA == APRTitulos.DATA,
            APRCapa.NUMERO == APRTitulos.NUMERO
        )
    ).join(
        Cedente,
        APRCapa.CEDENTE == Cedente.apelido
    ).outerjoin(
        ProdutoCedente,
        APRTitulos.id_produto == ProdutoCedente.Id
    ).outerjoin(
        Produto,
        ProdutoCedente.IdProdutoAtributo == Produto.Id
    ).filter(
        cast(APRCapa.DATA, Date) == target_date,
        APRTitulos.NFEChave.isnot(None),
        APRTitulos.NFEChave != ''
    ).distinct()

    results_with_nfe = query_with_nfe.all()
    print(f"Found {len(results_with_nfe)} records with NFEChave")

    # Products that don't need NFEChave validation (case-insensitive)
    products_no_nfechave_validation = [
        'CTE BMA FIDC',
        'CTE PRE-IMPRESSO BMA FIDC',
        'INTERCIA NFSE',
        'NF SERV. BMA SEC.',
        'CONTRATO',
        'COB SIMPLES NÃO PG 3ºS',
        'SEM NOTA',
        'CAPITAL DE GIRO NP',
        'RENEGOCIAÇÃO',
        'NOTA COMERCIAL',
        'MEIA NOTA',
        'NF SERV. PRE-IMPR. BMA FIDC',
        'NF SERV. BMA FIDC',
        'CCB',
        'CONVENCIONAL BMA FIDC',
        'CHEQUE',
    ]
    # Convert to uppercase for case-insensitive comparison
    products_no_nfechave_validation = [p.upper() for p in products_no_nfechave_validation]

    # Products that don't need ANY DUPLICATA validation (case-insensitive)
    products_no_duplicata_validation = [
        'CTE BMA FIDC',
        'CTE PRE-IMPRESSO BMA FIDC',
        'CONTRATO',
        'COB SIMPLES NÃO PG 3ºS',
        'SEM NOTA',
        'COB SIMPLES GARANTIA',
        'RENEGOCIAÇÃO',
        'NOTA COMERCIAL',
        'MEIA NOTA',
        'COBRANÇA SIMPLES',
        'CHEQUE',  # CHEQUE products don't need DUPLICATA validation
    ]
    # Convert to uppercase for case-insensitive comparison
    products_no_duplicata_validation = [p.upper() for p in products_no_duplicata_validation]

    # Products that don't need SEUNO validation (case-insensitive)
    products_no_seuno_validation = [
        'COMISSÁRIA',
        'ESCROW DEPÓSITO',
        'CHEQUE',
    ]
    # Convert to uppercase for case-insensitive comparison
    products_no_seuno_validation = [p.upper() for p in products_no_seuno_validation]

    # Query 2b: Records WITHOUT NFEChave - ALL products
    # Changed to OUTERJOIN to include ALL products, not just specific ones
    # Validation rules will be applied conditionally based on product type
    query_without_nfe = session.query(
        APRCapa.GERENTE,
        APRCapa.NUMERO.label('PROPOSTA'),
        APRCapa.DATA,
        APRCapa.CEDENTE,
        APRCapa.empresa,
        APRCapa.status_atual.label('STATUS'),
        Cedente.ramo.label('RAMO'),
        APRTitulos.SEUNO,
        APRTitulos.TITULO.label('DUPLICATA'),
        APRTitulos.NFEChave,
        Produto.Descritivo.label('PRODUTO'),
        Cedente.apelido
    ).join(
        APRTitulos,
        and_(
            APRCapa.DATA == APRTitulos.DATA,
            APRCapa.NUMERO == APRTitulos.NUMERO
        )
    ).join(
        Cedente,
        APRCapa.CEDENTE == Cedente.apelido
    ).outerjoin(
        ProdutoCedente,
        APRTitulos.id_produto == ProdutoCedente.Id
    ).outerjoin(
        Produto,
        ProdutoCedente.IdProdutoAtributo == Produto.Id
    ).filter(
        cast(APRCapa.DATA, Date) == target_date,
        or_(
            APRTitulos.NFEChave.is_(None),
            APRTitulos.NFEChave == ''
        )
    ).distinct()

    results_without_nfe = query_without_nfe.all()
    print(f"Found {len(results_without_nfe)} records without NFEChave (all products)")

    # Combine both result sets
    results = list(results_with_nfe) + list(results_without_nfe)
    print(f"Total records to validate: {len(results)}")

    # Step 3: Process each record and validate both DUPLICATA and SEUNO
    print("\nStep 3: Validating DUPLICATA and SEUNO...")

    validated_results = []

    for record in results:
        gerente = record.GERENTE
        proposta = record.PROPOSTA
        data = record.DATA
        cedente = record.CEDENTE
        empresa = record.empresa
        status = record.STATUS
        ramo = record.RAMO
        seuno = record.SEUNO
        duplicata = record.DUPLICATA
        nfe_chave = record.NFEChave
        produto = record.PRODUTO

        # Apply status filter if enabled
        if apply_status_filter:
            if not is_status_equal_or_further(status, "Aguardando Analista"):
                continue  # Skip records that don't meet the status requirement

        # Extract NFE from NFEChave (positions 26-34, 9 digits)
        nfe = None
        if nfe_chave and len(nfe_chave) >= 34:
            try:
                nfe = int(nfe_chave[25:34])  # 0-indexed, so position 26 is index 25
            except (ValueError, IndexError):
                nfe = None

        # Validate NFEChave FIRST (before DUPLICATA and SEUNO)
        # Skip validation for products that don't require it (case-insensitive)
        produto_upper = produto.upper() if produto else ""

        # DEFENSIVE CHECK: If PRODUTO is NULL/missing, skip ALL validations
        # We cannot determine validation rules without knowing the product type
        if not produto or not produto.strip():
            # Log warning and skip validation for this record
            print(f"  ⚠️  WARNING: Proposta {proposta} - Título with NULL/missing PRODUTO - skipping validation")
            print(f"     CEDENTE: {cedente}, DUPLICATA: {duplicata}, SEUNO: {seuno}")
            # Mark as valid since we can't validate without product info
            nfechave_valid = True
            nfechave_motivo = ""
            duplicata_valid = True
            duplicata_motivo = ""
            seuno_valid = True
            seuno_motivo = ""
            seuno_range = ""
            seuno_company = ""
            verification_digit = ""
        else:
            # PRODUTO is available - proceed with normal validation
            nfechave_valid = True
            nfechave_motivo = ""

            if produto_upper not in products_no_nfechave_validation:
                # This product requires NFEChave validation
                nfechave_valid = validate_nfechave(nfe_chave)
                if not nfechave_valid:
                    nfechave_motivo = get_invalid_nfechave_reason(produto or "PRODUTO")

            # Validate DUPLICATA (only if NFEChave is valid)
            # Skip validation for products that don't require it (case-insensitive)
            duplicata_valid = True
            duplicata_motivo = ""

            if nfechave_valid:  # Only validate DUPLICATA if NFEChave is valid
                if produto_upper in products_no_duplicata_validation:
                    # No DUPLICATA validation needed for these products
                    duplicata_valid = True
                    duplicata_motivo = ""
                elif nfe:
                    # NFE exists - use full validation
                    duplicata_valid = validate_duplicata_format(duplicata, nfe)
                    duplicata_motivo = "" if duplicata_valid else get_invalid_duplicata_reason(duplicata, nfe)
                else:
                    # No NFE - use simple validation (only check separator + sequential)
                    duplicata_valid = validate_duplicata_format_simple(duplicata)
                    duplicata_motivo = "" if duplicata_valid else get_invalid_duplicata_reason_simple(duplicata)

            # Validate SEUNO (only if NFEChave is valid and for cedentes with pre-impresso)
            # Skip validation for products that don't require it (case-insensitive)
            seuno_valid = True
            seuno_motivo = ""
            seuno_range = ""
            seuno_company = ""
            verification_digit = ""

            if nfechave_valid:  # Only validate SEUNO if NFEChave is valid
                if produto_upper in products_no_seuno_validation:
                    # No SEUNO validation needed for these products
                    seuno_valid = True
                    seuno_motivo = ""
                elif cedente in cedente_ranges:
                    # This cedente has pre-impresso ranges
                    ranges_list = cedente_ranges[cedente]

                    # NEW LOGIC: Try ALL ranges and accept if ANY match
                    # This fixes the issue where cedentes with multiple ranges (e.g., COMBRASIL with ranges 3 and 5061)
                    # would only validate against the first matching range, causing false negatives

                    verification_digit = calculate_verification_digit(seuno or '')
                    seuno_valid = False
                    seuno_motivo = ""
                    matched_range = None
                    all_validation_errors = []

                    # Try each range until we find a valid match
                    for company, range_val in ranges_list:
                        is_valid, motivo = validate_seuno(seuno, range_val, verification_digit)

                        if is_valid:
                            # Found a valid range! Mark as valid and record which range matched
                            seuno_valid = True
                            seuno_motivo = ""
                            matched_range = {'company': company, 'range': range_val}
                            seuno_company = company
                            seuno_range = range_val
                            break  # No need to check other ranges
                        else:
                            # Record the error for this range
                            all_validation_errors.append(f"Range {range_val} ({company}): {motivo}")

                    # If no range matched, combine all error messages
                    if not seuno_valid:
                        seuno_motivo = " | ".join(all_validation_errors)

        # Determine overall validity
        is_valid = nfechave_valid and duplicata_valid and seuno_valid

        # Combine motivos
        motivos = []
        if nfechave_motivo:
            motivos.append(nfechave_motivo)
        if duplicata_motivo:
            motivos.append(f"DUPLICATA: {duplicata_motivo}")
        if seuno_motivo:
            motivos.append(f"SEUNO: {seuno_motivo}")
        motivo_invalido = "; ".join(motivos) if motivos else ""

        validated_results.append({
            'GERENTE': gerente,
            'PROPOSTA': proposta,
            'DATA': data,
            'CEDENTE': cedente,
            'EMPRESA': empresa,
            'STATUS': status,
            'RAMO': ramo,  # Add RAMO field
            'PRODUTO': produto,  # Add PRODUTO field
            'SEUNO': seuno,
            'DUPLICATA': duplicata,
            'NFE': nfe,
            'SEUNO_COMPANY': seuno_company,
            'SEUNO_RANGE': seuno_range,
            'VERIFICATION_DIGIT': verification_digit,
            'NFECHAVE_VALID': nfechave_valid,
            'DUPLICATA_VALID': duplicata_valid,
            'SEUNO_VALID': seuno_valid,
            'IS_VALID': is_valid,
            'MOTIVO_INVALIDO': motivo_invalido
        })

    return validated_results


def query_valid_records_with_status_filter(session, target_date, invalid_record_keys):
    """
    Query valid records (excluding invalid ones) with status filter.
    Only includes records with status "Aguardando Analista" or further.

    Args:
        session: SQLAlchemy session
        target_date: Date to query
        invalid_record_keys: Set of (DATA, NUMERO, DUPLICATA) tuples to exclude

    Returns:
        List of valid record dictionaries
    """
    print("\n" + "=" * 140)
    print("QUERYING VALID RECORDS WITH STATUS FILTER")
    print("=" * 140)

    # Query 1: Records WITH NFEChave
    query_with_nfe = session.query(
        APRCapa.GERENTE,
        APRCapa.NUMERO.label('PROPOSTA'),
        APRCapa.DATA,
        APRCapa.CEDENTE,
        APRCapa.empresa,
        APRCapa.status_atual.label('STATUS'),
        APRCapa.QTD_APROVADOS,
        APRCapa.VLR_APROVADOS,
        Cedente.ramo.label('RAMO'),
        APRTitulos.SEUNO,
        APRTitulos.TITULO.label('DUPLICATA'),
        APRTitulos.NFEChave,
        APRTitulos.VALOR,
        Cedente.apelido
    ).join(
        APRTitulos,
        and_(
            APRCapa.DATA == APRTitulos.DATA,
            APRCapa.NUMERO == APRTitulos.NUMERO
        )
    ).join(
        Cedente,
        APRCapa.CEDENTE == Cedente.apelido
    ).filter(
        cast(APRCapa.DATA, Date) == target_date,
        APRTitulos.NFEChave.isnot(None),
        APRTitulos.NFEChave != ''
    ).distinct()

    results_with_nfe = query_with_nfe.all()
    print(f"Found {len(results_with_nfe)} records with NFEChave")

    # Query 2: Records WITHOUT NFEChave - ALL products
    # Changed to OUTERJOIN to include ALL products, not just specific ones
    # Validation rules will be applied conditionally based on product type
    query_without_nfe = session.query(
        APRCapa.GERENTE,
        APRCapa.NUMERO.label('PROPOSTA'),
        APRCapa.DATA,
        APRCapa.CEDENTE,
        APRCapa.empresa,
        APRCapa.status_atual.label('STATUS'),
        APRCapa.QTD_APROVADOS,
        APRCapa.VLR_APROVADOS,
        Cedente.ramo.label('RAMO'),
        APRTitulos.SEUNO,
        APRTitulos.TITULO.label('DUPLICATA'),
        APRTitulos.NFEChave,
        APRTitulos.VALOR,
        Cedente.apelido
    ).join(
        APRTitulos,
        and_(
            APRCapa.DATA == APRTitulos.DATA,
            APRCapa.NUMERO == APRTitulos.NUMERO
        )
    ).join(
        Cedente,
        APRCapa.CEDENTE == Cedente.apelido
    ).outerjoin(
        ProdutoCedente,
        APRTitulos.id_produto == ProdutoCedente.Id
    ).outerjoin(
        Produto,
        ProdutoCedente.IdProdutoAtributo == Produto.Id
    ).filter(
        cast(APRCapa.DATA, Date) == target_date,
        or_(
            APRTitulos.NFEChave.is_(None),
            APRTitulos.NFEChave == ''
        )
    ).distinct()

    results_without_nfe = query_without_nfe.all()
    print(f"Found {len(results_without_nfe)} records without NFEChave (all products)")

    # Combine both result sets
    results = list(results_with_nfe) + list(results_without_nfe)
    print(f"Total records to process: {len(results)}")

    # Filter out invalid records and apply status filter
    valid_records = []
    skipped_by_status = 0

    for record in results:
        # Create key to check if this record is invalid
        record_key = (record.DATA, record.PROPOSTA, record.DUPLICATA)

        # Skip if this record is in the invalid set
        if record_key in invalid_record_keys:
            continue

        # Apply status filter: only include records with status >= "Aguardando Analista"
        status = record.STATUS or ''
        if not is_status_equal_or_further(status, "Aguardando Analista"):
            skipped_by_status += 1
            continue

        valid_records.append({
            'GERENTE': record.GERENTE,
            'PROPOSTA': record.PROPOSTA,
            'DATA': record.DATA,
            'CEDENTE': record.CEDENTE,
            'EMPRESA': record.empresa,
            'STATUS': record.STATUS,
            'QTD_APROVADOS': int(record.QTD_APROVADOS) if record.QTD_APROVADOS else 0,
            'VLR_APROVADOS': float(record.VLR_APROVADOS) if record.VLR_APROVADOS else 0.0,
            'RAMO': record.RAMO,
            'SEUNO': record.SEUNO,
            'DUPLICATA': record.DUPLICATA,
            'VALOR': float(record.VALOR) if record.VALOR else 0.0
        })

    print(f"Found {len(valid_records)} valid records")
    print(f"Excluded {len(invalid_record_keys)} invalid records")
    print(f"Skipped {skipped_by_status} records due to status filter (status < 'Aguardando Analista')")

    return valid_records


# ============================================================================
# EXCEL EXPORT FUNCTIONS
# ============================================================================

def export_invalid_duplicata_to_excel(invalid_records, target_date):
    """
    Export invalid DUPLICATA records to Excel file.

    Args:
        invalid_records: List of invalid record dictionaries
        target_date: Date object for the report

    Returns:
        Path to the created Excel file
    """
    # Create output directory
    output_dir = Path(__file__).parent / 'duplicatas_invalidas'
    output_dir.mkdir(exist_ok=True)

    # Generate filename with timestamp
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f"duplicatas_invalidas_{target_date.strftime('%Y-%m-%d')}_{timestamp}.xlsx"
    filepath = output_dir / filename

    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "DUPLICATA Inválida"

    # Define headers
    headers = [
        'GERENTE', 'PROPOSTA', 'DATA', 'CEDENTE', 'EMPRESA', 'STATUS', 'RAMO',
        'DUPLICATA', 'NFE', 'MOTIVO_INVALIDO'
    ]

    # Write headers with formatting
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Write data rows
    for row_num, record in enumerate(invalid_records, 2):
        ws.cell(row=row_num, column=1, value=record.get('GERENTE'))
        ws.cell(row=row_num, column=2, value=record.get('PROPOSTA'))
        ws.cell(row=row_num, column=3, value=record.get('DATA'))
        ws.cell(row=row_num, column=4, value=record.get('CEDENTE'))
        ws.cell(row=row_num, column=5, value=record.get('EMPRESA'))
        ws.cell(row=row_num, column=6, value=record.get('STATUS'))
        ws.cell(row=row_num, column=7, value=record.get('RAMO'))
        ws.cell(row=row_num, column=8, value=record.get('DUPLICATA'))
        ws.cell(row=row_num, column=9, value=record.get('NFE'))
        ws.cell(row=row_num, column=10, value=record.get('MOTIVO_INVALIDO'))

    # Auto-adjust column widths
    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        max_length = len(header)
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=col_num, max_col=col_num):
            for cell in row:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_length + 2, 50)

    # Save workbook
    wb.save(filepath)
    print(f"\n✓ Excel file created: {filepath}")

    return filepath


def export_invalid_nfechave_to_excel(invalid_records, target_date):
    """
    Export invalid NFEChave records to Excel file.

    Args:
        invalid_records: List of invalid record dictionaries
        target_date: Date object for the report

    Returns:
        Path to the created Excel file
    """
    # Create output directory
    output_dir = Path(__file__).parent / 'nfechave_ausente'
    output_dir.mkdir(exist_ok=True)

    # Generate filename with timestamp
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f"nfechave_ausente_{target_date.strftime('%Y-%m-%d')}_{timestamp}.xlsx"
    filepath = output_dir / filename

    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "NFEChave Ausente"

    # Define headers
    headers = [
        'GERENTE', 'PROPOSTA', 'DATA', 'CEDENTE', 'EMPRESA', 'STATUS', 'RAMO',
        'PRODUTO', 'MOTIVO_INVALIDO'
    ]

    # Write headers with formatting
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Write data rows
    for row_num, record in enumerate(invalid_records, 2):
        ws.cell(row=row_num, column=1, value=record.get('GERENTE'))
        ws.cell(row=row_num, column=2, value=record.get('PROPOSTA'))
        ws.cell(row=row_num, column=3, value=record.get('DATA'))
        ws.cell(row=row_num, column=4, value=record.get('CEDENTE'))
        ws.cell(row=row_num, column=5, value=record.get('EMPRESA'))
        ws.cell(row=row_num, column=6, value=record.get('STATUS'))
        ws.cell(row=row_num, column=7, value=record.get('RAMO'))
        ws.cell(row=row_num, column=8, value=record.get('PRODUTO'))
        ws.cell(row=row_num, column=9, value=record.get('MOTIVO_INVALIDO'))

    # Auto-adjust column widths
    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        max_length = len(header)
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=col_num, max_col=col_num):
            for cell in row:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_length + 2, 50)

    # Save workbook
    wb.save(filepath)
    print(f"\n✓ Excel file created: {filepath}")

    return filepath


def export_invalid_seuno_to_excel(invalid_records, target_date):
    """
    Export invalid SEUNO records to Excel file.

    Args:
        invalid_records: List of invalid record dictionaries
        target_date: Date object for the report

    Returns:
        Path to the created Excel file
    """
    # Create output directory
    output_dir = Path(__file__).parent / 'seuno_invalidos'
    output_dir.mkdir(exist_ok=True)

    # Generate filename with timestamp
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f"seuno_invalidos_{target_date.strftime('%Y-%m-%d')}_{timestamp}.xlsx"
    filepath = output_dir / filename

    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "SEUNO Inválido"

    # Define headers
    headers = [
        'GERENTE', 'PROPOSTA', 'DATA', 'CEDENTE', 'EMPRESA', 'STATUS', 'RAMO',
        'SEUNO', 'DUPLICATA', 'SEUNO_COMPANY', 'SEUNO_RANGE',
        'VERIFICATION_DIGIT', 'MOTIVO_INVALIDO'
    ]

    # Write headers with formatting
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Write data rows
    for row_num, record in enumerate(invalid_records, 2):
        ws.cell(row=row_num, column=1, value=record.get('GERENTE'))
        ws.cell(row=row_num, column=2, value=record.get('PROPOSTA'))
        ws.cell(row=row_num, column=3, value=record.get('DATA'))
        ws.cell(row=row_num, column=4, value=record.get('CEDENTE'))
        ws.cell(row=row_num, column=5, value=record.get('EMPRESA'))
        ws.cell(row=row_num, column=6, value=record.get('STATUS'))
        ws.cell(row=row_num, column=7, value=record.get('RAMO'))
        ws.cell(row=row_num, column=8, value=record.get('SEUNO'))
        ws.cell(row=row_num, column=9, value=record.get('DUPLICATA'))
        ws.cell(row=row_num, column=10, value=record.get('SEUNO_COMPANY'))
        ws.cell(row=row_num, column=11, value=record.get('SEUNO_RANGE'))
        ws.cell(row=row_num, column=12, value=record.get('VERIFICATION_DIGIT'))
        ws.cell(row=row_num, column=13, value=record.get('MOTIVO_INVALIDO'))

    # Auto-adjust column widths
    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        max_length = len(header)
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=col_num, max_col=col_num):
            for cell in row:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_length + 2, 50)

    # Save workbook
    wb.save(filepath)
    print(f"\n✓ Excel file created: {filepath}")

    return filepath


def export_invalid_cheque_to_excel(invalid_records, target_date):
    """
    Export invalid CHEQUE records to Excel file.

    Args:
        invalid_records: List of invalid record dictionaries
        target_date: Date object for the report

    Returns:
        Path to the created Excel file
    """
    # Create output directory
    output_dir = Path(__file__).parent / 'cheque_invalidos'
    output_dir.mkdir(exist_ok=True)

    # Generate filename with timestamp
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f"cheque_invalidos_{target_date.strftime('%Y-%m-%d')}_{timestamp}.xlsx"
    filepath = output_dir / filename

    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "CHEQUE Inválido"

    # Define headers
    headers = [
        'GERENTE', 'PROPOSTA', 'DATA', 'CEDENTE', 'EMPRESA', 'STATUS', 'RAMO',
        'PRODUTO', 'QTD_TITULOS', 'MOTIVO_INVALIDO'
    ]

    # Write headers with formatting
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Write data rows
    for row_num, record in enumerate(invalid_records, 2):
        ws.cell(row=row_num, column=1, value=record.get('GERENTE'))
        ws.cell(row=row_num, column=2, value=record.get('PROPOSTA'))
        ws.cell(row=row_num, column=3, value=record.get('DATA'))
        ws.cell(row=row_num, column=4, value=record.get('CEDENTE'))
        ws.cell(row=row_num, column=5, value=record.get('EMPRESA'))
        ws.cell(row=row_num, column=6, value=record.get('STATUS'))
        ws.cell(row=row_num, column=7, value=record.get('RAMO'))
        ws.cell(row=row_num, column=8, value=record.get('PRODUTO'))
        ws.cell(row=row_num, column=9, value=record.get('QTD_TITULOS'))
        ws.cell(row=row_num, column=10, value=record.get('MOTIVO_INVALIDO'))

    # Auto-adjust column widths
    for col_letter in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J']:
        column = ws[col_letter]
        max_length = 0
        for cell in column:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_length + 2, 50)

    # Save workbook
    wb.save(filepath)
    print(f"\n✓ Excel file created: {filepath}")

    return filepath



def get_records_hash(records: list[dict], alert_type: str) -> str:
    """
    Generate a hash of the invalid records to track if they've been sent.

    Hash is based on unique (DATA, CEDENTE, PROPOSTA, MOTIVO) combinations,
    NOT individual DUPLICATA/SEUNO values. This ensures we send only ONE alert
    per (CEDENTE, PROPOSTA) combination per day, regardless of how many
    individual invalid records exist.

    Args:
        records: List of invalid records
        alert_type: Type of alert ("DUPLICATA" or "SEUNO")

    Returns:
        SHA256 hash of the records
    """
    # Group by unique (CEDENTE, PROPOSTA, MOTIVO) combinations
    # NOTE: DATA is NOT included to avoid duplicate alerts for the same proposal
    unique_combinations = set()

    for record in records:
        cedente = record.get('CEDENTE', '')
        proposta = record.get('PROPOSTA', '')
        motivo = record.get('MOTIVO_INVALIDO', '')

        # Create key based on (CEDENTE, PROPOSTA, MOTIVO) only
        # This ensures we send only ONE alert per unique combination per day
        key = f"{cedente}|{proposta}|{motivo}"
        unique_combinations.add(key)

    # Sort for consistency
    sorted_keys = sorted(unique_combinations)
    records_str = '|'.join(sorted_keys)

    return hashlib.sha256(records_str.encode()).hexdigest()


def get_tracking_file(alert_type: str) -> Path:
    """
    Get the path to the tracking file for today.

    Uses current working directory to ensure it works correctly when
    running as a PyInstaller binary.

    Args:
        alert_type: Type of alert ("DUPLICATA", "SEUNO", or "NFECHAVE")

    Returns:
        Path to tracking file
    """
    # Use current working directory instead of __file__ to work with PyInstaller
    base_dir = Path.cwd()

    if alert_type == "DUPLICATA":
        tracking_dir = base_dir / 'duplicatas_invalidas_tracking'
    elif alert_type == "SEUNO":
        tracking_dir = base_dir / 'seuno_invalidos_tracking'
    else:  # NFECHAVE
        tracking_dir = base_dir / 'nfechave_ausente_tracking'

    tracking_dir.mkdir(exist_ok=True)

    today = date.today().strftime('%Y-%m-%d')
    return tracking_dir / f'.tracking_nfechave_{today}.json' if alert_type == "NFECHAVE" else tracking_dir / f'.tracking_{today}.json'


def get_new_combinations(records: list[dict], alert_type: str) -> list[dict]:
    """
    Filter records to return only NEW combinations that haven't been alerted today.

    This ensures we only alert about new (CEDENTE, PROPOSTA, MOTIVO) combinations,
    not re-alert about combinations that were already sent earlier today.

    Args:
        records: List of invalid records
        alert_type: Type of alert ("DUPLICATA" or "SEUNO")

    Returns:
        List of records with only NEW combinations that haven't been alerted yet
    """
    tracking_file = get_tracking_file(alert_type)

    # If no tracking file exists, all records are new
    if not tracking_file.exists():
        return records

    try:
        with open(tracking_file, 'r') as f:
            tracking_data = json.load(f)

        # Get the set of already-alerted combinations
        alerted_combinations = set(tracking_data.get('alerted_combinations', []))

        # Filter to only NEW combinations
        new_records = []
        for record in records:
            cedente = record.get('CEDENTE', '')
            proposta = record.get('PROPOSTA', '')
            motivo = record.get('MOTIVO_INVALIDO', '')

            combination_key = f"{cedente}|{proposta}|{motivo}"

            # Only include if this combination hasn't been alerted yet
            if combination_key not in alerted_combinations:
                new_records.append(record)

        return new_records

    except Exception as e:
        print(f"Warning: Error reading tracking file: {e}")
        # On error, return all records to be safe
        return records


def mark_as_sent(records: list[dict], alert_type: str):
    """
    Mark these record combinations as sent in the tracking file.

    Appends new combinations to the existing tracking file instead of replacing.
    This allows us to track individual combinations throughout the day.

    Args:
        records: List of invalid records that were just alerted
        alert_type: Type of alert ("DUPLICATA" or "SEUNO")
    """
    tracking_file = get_tracking_file(alert_type)

    # Read existing tracking data
    alerted_combinations = []
    if tracking_file.exists():
        try:
            with open(tracking_file, 'r') as f:
                tracking_data = json.load(f)
                alerted_combinations = tracking_data.get('alerted_combinations', [])
        except Exception as e:
            print(f"Warning: Error reading existing tracking file: {e}")

    # Add new combinations
    for record in records:
        cedente = record.get('CEDENTE', '')
        proposta = record.get('PROPOSTA', '')
        motivo = record.get('MOTIVO_INVALIDO', '')

        combination_key = f"{cedente}|{proposta}|{motivo}"

        # Avoid duplicates in the tracking file
        if combination_key not in alerted_combinations:
            alerted_combinations.append(combination_key)

    # Save updated tracking data
    tracking_data = {
        'alerted_combinations': alerted_combinations,
        'last_updated': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        'alert_type': alert_type
    }

    try:
        with open(tracking_file, 'w') as f:
            json.dump(tracking_data, f, indent=2)
        print(f"✓ Marked {len(records)} new {alert_type} combinations as sent (total tracked: {len(alerted_combinations)})")
    except Exception as e:
        print(f"Warning: Error writing tracking file: {e}")


def send_slack_notification(excel_file, invalid_count, target_date, alert_type="DUPLICATA", invalid_records=None, dry_run=True):
    """
    Send Slack notification with Excel file attachment using files.uploadV2 API.

    Args:
        excel_file: Path to Excel file
        invalid_count: Number of invalid records
        target_date: Date object for the report
        alert_type: Type of alert ("DUPLICATA", "SEUNO", or "NFECHAVE")
        invalid_records: List of invalid record dictionaries
        dry_run: If True, don't actually send to Slack
    """
    try:
        # Filter to only NEW combinations that haven't been alerted yet today
        if invalid_records:
            original_count = len(invalid_records)
            new_records = get_new_combinations(invalid_records, alert_type)

            if not new_records:
                print(f"\n⏭️  {alert_type} notification skipped - all combinations already alerted today")
                print(f"   Total records: {original_count}, New combinations: 0")
                return

            # Update the records list to only include new combinations
            invalid_records = new_records
            invalid_count = len(new_records)
            print(f"\n📊 Found {invalid_count} NEW {alert_type} combinations to alert (filtered from {original_count} total)")

            # Regenerate Excel file with only NEW records
            if alert_type == "DUPLICATA":
                excel_file = export_invalid_duplicata_to_excel(invalid_records, target_date)
            elif alert_type == "SEUNO":
                excel_file = export_invalid_seuno_to_excel(invalid_records, target_date)
            else:  # NFECHAVE
                excel_file = export_invalid_nfechave_to_excel(invalid_records, target_date)

        if dry_run:
            print(f"\n[DRY RUN] Would upload {alert_type} alert to Slack:")
            print(f"  File: {excel_file}")
            print(f"  Count: {invalid_count} NEW invalid combinations")
            print(f"  Date: {target_date.strftime('%d/%m/%Y')}")
            print(f"  Note: Marking as sent to prevent duplicate notifications in subsequent runs")
            # Mark as sent even in dry-run mode to prevent duplicate alerts
            if invalid_records:
                mark_as_sent(invalid_records, alert_type)
            return

        slack_config = load_slack_config()
        bot_token = slack_config.get('slack', {}).get('bot_token')
        channel = slack_config.get('slack', {}).get('channel', '#alerts')

        if not bot_token:
            print("⚠ Slack bot token not configured. Skipping notification.")
            return

        # Upload file to Slack using the new files.uploadV2 API
        print(f"\nUploading to Slack channel: {channel}")

        headers = {"Authorization": f"Bearer {bot_token}"}

        # Build message with list of cedentes
        if alert_type == "SEUNO":
            message_title = "⚠️ SEUNO Inválidos - Relatório de Pré-Impresso"
        elif alert_type == "NFECHAVE":
            message_title = "⚠️ NFEChave Ausente - Chaves de NF Faltando"
        else:
            message_title = "🚨 DUPLICATAS com formato inválido"

        message_lines = [f"*{message_title}*\n"]

        # Add cedente details - only unique combinations
        if invalid_records:
            # Use a set to track unique combinations of (CEDENTE, PROPOSTA, MOTIVO)
            unique_errors = set()
            for record in invalid_records:
                cedente = record.get('CEDENTE', 'N/A')
                proposta = record.get('PROPOSTA', 'N/A')
                motivo = record.get('MOTIVO_INVALIDO', 'N/A')
                # Remove the "SEUNO: " or "DUPLICATA: " prefix from motivo if present
                if alert_type == "SEUNO" and motivo.startswith("SEUNO: "):
                    motivo = motivo[7:]
                elif alert_type == "DUPLICATA" and motivo.startswith("DUPLICATA: "):
                    motivo = motivo[11:]
                # For NFECHAVE, motivo is already in the correct format
                unique_errors.add((cedente, proposta, motivo))

            # Sort and add to message
            for cedente, proposta, motivo in sorted(unique_errors):
                message_lines.append(f"{cedente}, {proposta}, {motivo}")

        message = "\n".join(message_lines)

        if alert_type == "DUPLICATA":
            alert_title = "DUPLICATA Inválida"
        elif alert_type == "SEUNO":
            alert_title = "SEUNO Inválido"
        else:  # NFECHAVE
            alert_title = "NFEChave Ausente"
        filename = f'{alert_title}_{target_date.strftime("%Y-%m-%d")}.xlsx'

        # Step 1: Post the message first using chat.postMessage
        message_response = requests.post(
            "https://slack.com/api/chat.postMessage",
            headers={**headers, "Content-Type": "application/json"},
            json={
                "channel": channel,
                "text": message
            }
        )
        message_result = message_response.json()

        if not message_result.get('ok'):
            print(f"⚠ Failed to post message: {message_result.get('error')}")
            # Continue anyway to try uploading the file

        # Step 2: Upload file using files.uploadV2 workflow
        # First, get upload URL
        file_size = Path(excel_file).stat().st_size

        upload_url_response = requests.post(
            "https://slack.com/api/files.getUploadURLExternal",
            headers=headers,
            data={
                "filename": filename,
                "length": file_size
            }
        )
        upload_url_result = upload_url_response.json()

        if not upload_url_result.get('ok'):
            error_msg = upload_url_result.get('error', 'unknown')
            print(f"⚠ Failed to get upload URL: {error_msg}")
            print(f"   Response: {upload_url_result}")
            return

        upload_url = upload_url_result['upload_url']
        file_id = upload_url_result['file_id']

        # Step 3: Upload file to the URL
        with open(excel_file, 'rb') as f:
            upload_response = requests.post(upload_url, data=f.read())

        if upload_response.status_code != 200:
            print(f"⚠ File upload failed with status code: {upload_response.status_code}")
            print(f"   Response: {upload_response.text}")
            return

        # Step 4: Complete the upload
        complete_response = requests.post(
            "https://slack.com/api/files.completeUploadExternal",
            headers=headers,
            data={
                "files": json.dumps([{
                    "id": file_id,
                    "title": f'{alert_title} - {target_date.strftime("%Y-%m-%d")}'
                }]),
                "channel_id": channel
            }
        )
        complete_result = complete_response.json()

        if complete_result.get('ok'):
            print("✓ Slack notification sent successfully!")
            # Mark as sent to prevent duplicate notifications
            if invalid_records:
                mark_as_sent(invalid_records, alert_type)
        else:
            print(f"⚠ Slack upload completion failed: {complete_result.get('error')}")
            print(f"   Response: {complete_result}")

    except Exception as e:
        print(f"⚠ Error sending Slack notification: {e}")
        import traceback
        traceback.print_exc()


def cleanup_old_tracking_files(alert_type="DUPLICATA", days_to_keep=7):
    """
    Remove old tracking files, keeping only recent ones.

    Args:
        alert_type: Type of alert ("DUPLICATA", "SEUNO", or "NFECHAVE")
        days_to_keep: Number of days of tracking files to keep (default: 7)
    """
    # Use current working directory instead of __file__ to work with PyInstaller
    base_dir = Path.cwd()

    if alert_type == "DUPLICATA":
        tracking_dir = base_dir / 'duplicatas_invalidas_tracking'
    elif alert_type == "SEUNO":
        tracking_dir = base_dir / 'seuno_invalidos_tracking'
    else:  # NFECHAVE
        tracking_dir = base_dir / 'nfechave_ausente_tracking'

    if not tracking_dir.exists():
        return

    # Calculate cutoff date
    cutoff_date = date.today() - timedelta(days=days_to_keep)

    pattern = '.tracking_nfechave_*.json' if alert_type == "NFECHAVE" else '.tracking_*.json'
    for file in tracking_dir.glob(pattern):
        try:
            # Extract date from filename: .tracking_YYYY-MM-DD.json or .tracking_nfechave_YYYY-MM-DD.json
            if alert_type == "NFECHAVE":
                date_str = file.stem.replace('.tracking_nfechave_', '')
            else:
                date_str = file.stem.replace('.tracking_', '')
            file_date = datetime.strptime(date_str, '%Y-%m-%d').date()

            if file_date < cutoff_date:
                file.unlink()
                print(f"Removed old tracking file: {file.name}")
        except Exception as e:
            print(f"Warning: Could not process tracking file {file.name}: {e}")


def cleanup_old_files(current_date, alert_type="DUPLICATA"):
    """
    Remove old Excel files, keeping only files from the current report date.

    Args:
        current_date: Date object for the current report
        alert_type: Type of alert ("DUPLICATA", "SEUNO", or "NFECHAVE")
    """
    if alert_type == "DUPLICATA":
        output_dir = Path(__file__).parent / 'duplicatas_invalidas'
        file_pattern = 'duplicatas_invalidas_*.xlsx'
    elif alert_type == "SEUNO":
        output_dir = Path(__file__).parent / 'seuno_invalidos'
        file_pattern = 'seuno_invalidos_*.xlsx'
    else:  # NFECHAVE
        output_dir = Path(__file__).parent / 'nfechave_ausente'
        file_pattern = 'nfechave_ausente_*.xlsx'

    if not output_dir.exists():
        return

    current_date_str = current_date.strftime('%Y-%m-%d')

    for file in output_dir.glob(file_pattern):
        # Check if file is from a different date
        if current_date_str not in file.name:
            try:
                file.unlink()
                print(f"Removed old file: {file.name}")
            except Exception as e:
                print(f"Warning: Could not remove {file.name}: {e}")

    # Also cleanup old tracking files
    cleanup_old_tracking_files(alert_type)


# ============================================================================
# DISPLAY FUNCTIONS
# ============================================================================

def print_summary(validated_results):
    """
    Print summary of validation results.

    Args:
        validated_results: List of all validated records
    """
    valid_records = [r for r in validated_results if r['IS_VALID']]
    invalid_records = [r for r in validated_results if not r['IS_VALID']]

    duplicata_invalid = [r for r in validated_results if not r['DUPLICATA_VALID']]
    seuno_invalid = [r for r in validated_results if not r['SEUNO_VALID']]
    both_invalid = [r for r in validated_results if not r['DUPLICATA_VALID'] and not r['SEUNO_VALID']]

    print("\n" + "=" * 140)
    print("VALIDATION SUMMARY (Status >= 'Aguardando Analista')")
    print("=" * 140)
    print(f"Total records processed: {len(validated_results)}")
    print(f"Valid records: {len(valid_records)}")
    print(f"Invalid records: {len(invalid_records)}")
    print(f"  - DUPLICATA invalid: {len(duplicata_invalid)}")
    print(f"  - SEUNO invalid: {len(seuno_invalid)}")
    print(f"  - Both invalid: {len(both_invalid)}")
    print("=" * 140)
    print("Note: Records with status < 'Aguardando Analista' are excluded from this report")
    print("=" * 140)

    # Group invalid records by motivo
    if invalid_records:
        print("\nINVALID RECORDS BY REASON:")
        print("-" * 140)
        motivo_counts = {}
        for r in invalid_records:
            motivo = r['MOTIVO_INVALIDO']
            motivo_counts[motivo] = motivo_counts.get(motivo, 0) + 1

        for motivo, count in sorted(motivo_counts.items(), key=lambda x: x[1], reverse=True):
            print(f"  [{count:4d} records] {motivo}")
        print("-" * 140)

    # Group by STATUS
    print("\nRECORDS BY STATUS:")
    print("-" * 140)
    status_counts = {}
    for r in validated_results:
        status = r.get('STATUS', 'N/A')
        status_counts[status] = status_counts.get(status, 0) + 1

    for status, count in sorted(status_counts.items(), key=lambda x: x[1], reverse=True):
        print(f"  {status}: {count} records")
    print("-" * 140)


def print_invalid_records_table(invalid_records, max_rows=50, title="INVALID RECORDS"):
    """
    Print table of invalid records.

    Args:
        invalid_records: List of invalid record dictionaries
        max_rows: Maximum number of rows to display
        title: Title for the table
    """
    if not invalid_records:
        return

    print("\n" + "=" * 140)
    print(f"{title} (showing first {min(len(invalid_records), max_rows)} of {len(invalid_records)})")
    print("=" * 140)
    print(f"{'#':<5} {'GERENTE':<15} {'PROPOSTA':<10} {'CEDENTE':<20} {'STATUS':<15} {'DUPLICATA':<15} {'SEUNO':<15} {'MOTIVO':<40}")
    print("-" * 140)

    for i, record in enumerate(invalid_records[:max_rows], 1):
        gerente = str(record.get('GERENTE', ''))[:14]
        proposta = str(record.get('PROPOSTA', ''))
        cedente = str(record.get('CEDENTE', ''))[:19]
        status = str(record.get('STATUS', ''))[:14]
        duplicata = str(record.get('DUPLICATA', ''))[:14]
        seuno = str(record.get('SEUNO', ''))[:14]
        motivo = str(record.get('MOTIVO_INVALIDO', ''))[:39]

        print(f"{i:<5} {gerente:<15} {proposta:<10} {cedente:<20} {status:<15} {duplicata:<15} {seuno:<15} {motivo:<40}")

    if len(invalid_records) > max_rows:
        print(f"\n... and {len(invalid_records) - max_rows} more invalid records")

    print("-" * 140)


def print_valid_records_table(valid_records, max_rows=100):
    """
    Print table of valid records in human-readable format.

    Args:
        valid_records: List of valid record dictionaries
        max_rows: Maximum number of rows to display
    """
    if not valid_records:
        print("\n✓ No valid records found!")
        return

    print("\n" + "=" * 200)
    print(f"VALID RECORDS - READY FOR PROCESSING (showing first {min(len(valid_records), max_rows)} of {len(valid_records)})")
    print("=" * 200)
    print(f"{'#':<5} {'DATA':<12} {'GERENTE':<12} {'PROP':<6} {'CEDENTE':<20} {'RAMO':<18} {'STATUS':<20} {'QTD':>5} {'VLR_APROV':>15} {'VALOR':>15}")
    print("-" * 200)

    total_valor = 0.0
    total_vlr_aprovados = 0.0
    total_qty_aprovados = 0

    for i, record in enumerate(valid_records[:max_rows], 1):
        data = record.get('DATA')
        data_str = data.strftime('%Y-%m-%d') if data else 'N/A'
        gerente = str(record.get('GERENTE', ''))[:11]
        proposta = str(record.get('PROPOSTA', ''))
        cedente = str(record.get('CEDENTE', ''))[:19]
        ramo = str(record.get('RAMO', ''))[:17]
        status = str(record.get('STATUS', ''))[:19]
        qty_aprovados = record.get('QTD_APROVADOS', 0)
        vlr_aprovados = record.get('VLR_APROVADOS', 0.0)
        valor = record.get('VALOR', 0.0)

        total_valor += valor
        total_vlr_aprovados += vlr_aprovados
        total_qty_aprovados += qty_aprovados

        vlr_aprovados_str = f"R$ {vlr_aprovados:,.2f}"
        valor_str = f"R$ {valor:,.2f}"

        print(f"{i:<5} {data_str:<12} {gerente:<12} {proposta:<6} {cedente:<20} {ramo:<18} {status:<20} {qty_aprovados:>5} {vlr_aprovados_str:>15} {valor_str:>15}")

    if len(valid_records) > max_rows:
        print(f"\n... and {len(valid_records) - max_rows} more valid records")
        # Calculate total for all records
        total_valor = sum(r.get('VALOR', 0.0) for r in valid_records)
        total_vlr_aprovados = sum(r.get('VLR_APROVADOS', 0.0) for r in valid_records)
        total_qty_aprovados = sum(r.get('QTD_APROVADOS', 0) for r in valid_records)

    print("-" * 200)
    print(f"{'TOTAL:':>99} {total_qty_aprovados:>5} R$ {total_vlr_aprovados:>13,.2f} R$ {total_valor:>13,.2f}")
    print("=" * 200)

    # Group by status
    print("\nVALID RECORDS BY STATUS:")
    print("-" * 200)
    status_counts = {}
    status_values = {}
    status_vlr_aprovados = {}
    status_qty_aprovados = {}

    for r in valid_records:
        status = r.get('STATUS', 'N/A')
        status_counts[status] = status_counts.get(status, 0) + 1
        status_values[status] = status_values.get(status, 0.0) + r.get('VALOR', 0.0)
        status_vlr_aprovados[status] = status_vlr_aprovados.get(status, 0.0) + r.get('VLR_APROVADOS', 0.0)
        status_qty_aprovados[status] = status_qty_aprovados.get(status, 0) + r.get('QTD_APROVADOS', 0)

    print(f"  {'STATUS':<30} {'COUNT':>8} {'QTD_APROV':>10} {'VLR_APROVADOS':>20} {'VALOR_TITULOS':>20}")
    print("-" * 200)

    for status in sorted(status_counts.keys(), key=lambda x: status_counts[x], reverse=True):
        count = status_counts[status]
        qty = status_qty_aprovados[status]
        vlr_aprov = status_vlr_aprovados[status]
        value = status_values[status]
        print(f"  {status:<30} {count:>8} {qty:>10} R$ {vlr_aprov:>17,.2f} R$ {value:>17,.2f}")
    print("-" * 200)


# ============================================================================
# MAIN FUNCTION
# ============================================================================

def main():
    """Main function."""
    try:
        # Track execution start time
        start_time = datetime.now()

        # Check for dry-run flag (default is now production mode)
        dry_run = '--dry-run' in sys.argv  # Only dry-run if explicitly requested

        if dry_run:
            print("=" * 140)
            print("🔍 DRY RUN MODE - No Slack notifications will be sent, no database updates")
            print("=" * 140)
        else:
            print("=" * 140)
            print("🚀 PRODUCTION MODE - Slack notifications will be sent, database will be updated")
            print("=" * 140)

        print("=" * 140)
        print("APR INVALID RECORDS CHECKER - NFECHAVE, DUPLICATA & SEUNO with STATUS")
        print("=" * 140)
        print("Status Filter: Only processing records with status >= 'Aguardando Analista'")
        print("=" * 140)
        print()

        # Get target date (use most recent weekday)
        target_date = get_most_recent_weekday()

        # Create database connection
        config = load_config()
        connection_string = create_connection_string(config)

        print(f"Connecting to database...")
        print(f"Server: {config['databases']['mssql']['server']}")
        print(f"Database: {config['databases']['mssql']['scheme']}")

        engine = create_engine(connection_string)
        Session = sessionmaker(bind=engine)
        session = Session()

        # Create MariaDB session for monitoring (if not dry-run)
        mariadb_session = None
        if not dry_run:
            mariadb_session = create_mariadb_session(config)

        # Query and validate records (with status filter enabled)
        validated_results = query_apr_invalidos_with_status(session, target_date, apply_status_filter=True)

        # ========================================================================
        # TRACK ALL STATUS CHANGES FROM MSSQL SOURCE
        # ========================================================================
        # This must be done BEFORE validation to capture ALL status changes
        # including for proposals that may be invalid
        if not dry_run and mariadb_session:
            print("\n" + "=" * 140)
            print("TRACKING STATUS CHANGES FROM MSSQL SOURCE")
            print("=" * 140)
            print("Querying ALL proposals from MSSQL to detect status changes...")

            try:
                status_changes = track_all_status_changes_from_source(
                    mssql_session=session,
                    mariadb_session=mariadb_session,
                    target_date=target_date,
                    change_source='SYSTEM'
                )
                print(f"✓ Tracked {status_changes} status changes (including value changes)")
            except Exception as e:
                print(f"⚠️  Warning: Could not track status changes: {e}")
                import traceback
                traceback.print_exc()

        # ========================================================================
        # CHEQUE VALIDATION: Check proposals with ONLY CHEQUE products
        # ========================================================================
        invalid_cheque, validated_results = validate_cheque_proposals(session, validated_results, target_date)

        # Separate by validation type
        invalid_nfechave = [r for r in validated_results if not r['NFECHAVE_VALID']]
        invalid_duplicata = [r for r in validated_results if not r['DUPLICATA_VALID'] and r['NFECHAVE_VALID']]
        invalid_seuno = [r for r in validated_results if not r['SEUNO_VALID'] and r['NFECHAVE_VALID']]
        all_invalid = [r for r in validated_results if not r['IS_VALID']]

        # Print summary
        print_summary(validated_results)

        # ========================================================================
        # STEP -1: Handle Invalid CHEQUE Records (BEFORE all other validations)
        # ========================================================================
        if invalid_cheque:
            print("\n" + "=" * 140)
            print("STEP -1: PROCESSING INVALID CHEQUE RECORDS (Quantity > 100)")
            print("=" * 140)

            # Print table
            print_invalid_records_table(invalid_cheque, title="INVALID CHEQUE RECORDS (QTD > 100)")

            # Export to Excel
            excel_file = export_invalid_cheque_to_excel(invalid_cheque, target_date)

            # Send Slack notification
            send_slack_notification(excel_file, len(invalid_cheque), target_date,
                                   alert_type="CHEQUE", invalid_records=invalid_cheque, dry_run=dry_run)

            # Log invalid records to monitoring database
            if not dry_run and mariadb_session:
                log_invalid_records(mariadb_session, invalid_cheque, 'CHEQUE',
                                   alerted_count=len(invalid_cheque))
                print(f"  ✓ Logged {len(invalid_cheque)} invalid CHEQUE records to monitoring database")

            # Cleanup old files
            cleanup_old_files(target_date, alert_type="CHEQUE")
        else:
            print("\n✓ No invalid CHEQUE records found (all CHEQUE-only proposals have <= 100 titles)!")

        # ========================================================================
        # STEP 0: Handle Invalid NFEChave Records (BEFORE DUPLICATA and SEUNO)
        # ========================================================================
        if invalid_nfechave:
            print("\n" + "=" * 140)
            print("STEP 0: PROCESSING INVALID NFECHAVE RECORDS")
            print("=" * 140)

            # Print table
            print_invalid_records_table(invalid_nfechave, title="INVALID NFECHAVE RECORDS")

            # Export to Excel
            excel_file = export_invalid_nfechave_to_excel(invalid_nfechave, target_date)

            # Send Slack notification
            send_slack_notification(excel_file, len(invalid_nfechave), target_date,
                                   alert_type="NFECHAVE", invalid_records=invalid_nfechave, dry_run=dry_run)

            # Log invalid records to monitoring database
            if not dry_run and mariadb_session:
                log_invalid_records(mariadb_session, invalid_nfechave, 'NFECHAVE',
                                   alerted_count=len(invalid_nfechave))
                print(f"  ✓ Logged {len(invalid_nfechave)} invalid NFEChave records to monitoring database")

            # Cleanup old files
            cleanup_old_files(target_date, alert_type="NFECHAVE")
        else:
            print("\n✓ No invalid NFEChave records found!")

        # ========================================================================
        # STEP 1: Handle Invalid DUPLICATA Records
        # ========================================================================
        if invalid_duplicata:
            print("\n" + "=" * 140)
            print("STEP 1: PROCESSING INVALID DUPLICATA RECORDS")
            print("=" * 140)

            # Print table
            print_invalid_records_table(invalid_duplicata, title="INVALID DUPLICATA RECORDS")

            # Export to Excel
            excel_file = export_invalid_duplicata_to_excel(invalid_duplicata, target_date)

            # Send Slack notification
            send_slack_notification(excel_file, len(invalid_duplicata), target_date,
                                   alert_type="DUPLICATA", invalid_records=invalid_duplicata, dry_run=dry_run)

            # Log invalid records to monitoring database
            if not dry_run and mariadb_session:
                log_invalid_records(mariadb_session, invalid_duplicata, 'DUPLICATA',
                                   alerted_count=len(invalid_duplicata))
                print(f"  ✓ Logged {len(invalid_duplicata)} invalid DUPLICATA records to monitoring database")

            # Cleanup old files
            cleanup_old_files(target_date, alert_type="DUPLICATA")
        else:
            print("\n✓ No invalid DUPLICATA records found!")

        # ========================================================================
        # STEP 2: Handle Invalid SEUNO Records
        # ========================================================================
        if invalid_seuno:
            print("\n" + "=" * 140)
            print("STEP 2: PROCESSING INVALID SEUNO RECORDS")
            print("=" * 140)

            # Print table
            print_invalid_records_table(invalid_seuno, title="INVALID SEUNO RECORDS")

            # Export to Excel
            excel_file = export_invalid_seuno_to_excel(invalid_seuno, target_date)

            # Send Slack notification
            send_slack_notification(excel_file, len(invalid_seuno), target_date,
                                   alert_type="SEUNO", invalid_records=invalid_seuno, dry_run=dry_run)

            # Log invalid records to monitoring database
            if not dry_run and mariadb_session:
                log_invalid_records(mariadb_session, invalid_seuno, 'SEUNO',
                                   alerted_count=len(invalid_seuno))
                print(f"  ✓ Logged {len(invalid_seuno)} invalid SEUNO records to monitoring database")

            # Cleanup old files
            cleanup_old_files(target_date, alert_type="SEUNO")
        else:
            print("\n✓ No invalid SEUNO records found!")

        # ========================================================================
        # STEP 3: Query and Display Valid Records
        # ========================================================================
        print("\n" + "=" * 140)
        print("STEP 3: QUERYING VALID RECORDS WITH STATUS FILTER")
        print("=" * 140)

        # Create set of invalid PROPOSALS to exclude (proposal-level exclusion)
        # If ANY record in a proposal has invalid NFEChave, DUPLICATA, or SEUNO,
        # exclude the ENTIRE proposal from apr_valid_records table
        invalid_proposals = set()
        for r in all_invalid:
            invalid_proposals.add((r['DATA'], r['PROPOSTA']))

        print(f"\n⚠️  Excluding {len(invalid_proposals)} proposals with invalid records from apr_valid_records table")
        print(f"   (Proposals with ANY invalid NFEChave, DUPLICATA, or SEUNO record)")

        # Create set of invalid record keys to exclude (for backward compatibility)
        invalid_keys = set()
        for r in all_invalid:
            invalid_keys.add((r['DATA'], r['PROPOSTA'], r['DUPLICATA']))

        # Query valid records
        valid_records = query_valid_records_with_status_filter(session, target_date, invalid_keys)

        # ADDITIONAL FILTER: Exclude entire proposals that have ANY invalid record
        # This ensures proposal-level exclusion as requested
        valid_records_filtered = []
        for record in valid_records:
            proposal_key = (record['DATA'], record['PROPOSTA'])
            if proposal_key not in invalid_proposals:
                valid_records_filtered.append(record)

        excluded_count = len(valid_records) - len(valid_records_filtered)
        if excluded_count > 0:
            print(f"   Filtered out {excluded_count} additional valid records from proposals with invalid records")

        valid_records = valid_records_filtered

        # Display valid records
        print_valid_records_table(valid_records)

        # ========================================================================
        # STEP 4: Store Valid Records in MariaDB
        # ========================================================================
        print("\n" + "=" * 140)
        print("STEP 4: STORING VALID RECORDS IN MARIADB")
        print("=" * 140)

        try:
            # Connect to MariaDB
            mariadb_conn = create_mariadb_connection(config)

            # Store valid records
            inserted, updated = store_valid_records_to_mariadb(valid_records, mariadb_conn, dry_run)

            if not dry_run:
                print(f"\n✓ MariaDB storage completed:")
                print(f"  - Inserted: {inserted} new records")
                print(f"  - Updated: {updated} existing records")
                print(f"  - Total processed: {inserted + updated} records")

                # Track products for each proposal
                if mariadb_session:
                    print(f"\n📊 Tracking product types for proposals...")
                    products_tracked = track_proposal_products(mariadb_session, session, target_date)
                    print(f"  ✓ Tracked {products_tracked} product entries")

                # Note: Status changes are now tracked at the beginning of the program
                # using track_all_status_changes_from_source() which queries MSSQL directly

            # Close MariaDB connection
            mariadb_conn.close()

        except Exception as e:
            print(f"\n⚠ Warning: Could not store to MariaDB: {e}")
            print("  Continuing without MariaDB storage...")

        # Close MSSQL session
        session.close()

        # ========================================================================
        # STEP 5: Finalize Monitoring (if not dry-run)
        # ========================================================================
        if not dry_run and mariadb_session:
            print("\n" + "=" * 140)
            print("STEP 5: FINALIZING MONITORING")
            print("=" * 140)

            # Calculate execution time
            execution_time = int((datetime.now() - start_time).total_seconds())

            # Prepare execution statistics
            stats = {
                'total_records_queried': len(validated_results),
                'valid_records': len(valid_records),
                'invalid_nfechave': len(invalid_nfechave),
                'invalid_duplicata': len(invalid_duplicata),
                'invalid_seuno': len(invalid_seuno),
                'invalid_cheque': len(invalid_cheque) if 'invalid_cheque' in locals() else 0,
                'records_stored': inserted + updated if 'inserted' in locals() else 0,
                'alerts_sent_nfechave': len(invalid_nfechave) if invalid_nfechave else 0,
                'alerts_sent_duplicata': len(invalid_duplicata) if invalid_duplicata else 0,
                'alerts_sent_seuno': len(invalid_seuno) if invalid_seuno else 0,
                'alerts_sent_cheque': len(invalid_cheque) if 'invalid_cheque' in locals() and invalid_cheque else 0
            }

            # Log program execution
            print(f"\n📝 Logging program execution...")
            log_program_execution(mariadb_session, target_date, stats, execution_time,
                                 run_mode='PRODUCTION')
            print(f"  ✓ Execution logged (runtime: {execution_time}s)")

            # Update daily summary
            print(f"\n📊 Updating daily summary...")
            update_daily_summary(mariadb_session, target_date)
            print(f"  ✓ Daily summary updated")

            # Auto-resolve invalid records that are now valid
            print(f"\n🔄 Checking for auto-resolved records...")
            resolved = auto_resolve_invalid_records(mariadb_session, target_date)
            if resolved > 0:
                print(f"  ✓ Auto-resolved {resolved} invalid records")
            else:
                print(f"  ℹ️  No records to auto-resolve")

            # Cleanup old monitoring data (30-day retention)
            print(f"\n🧹 Cleaning up old monitoring data (30-day retention)...")
            deleted = cleanup_old_monitoring_data(mariadb_session, retention_days=30)
            total_deleted = sum(deleted.values())
            if total_deleted > 0:
                print(f"  ✓ Cleaned up {total_deleted} old records:")
                for table, count in deleted.items():
                    if count > 0:
                        print(f"    - {table}: {count} records")
            else:
                print(f"  ℹ️  No old data to clean up")

            # Close MariaDB session
            mariadb_session.close()
            print(f"\n✓ Monitoring finalized")

        print("\n" + "=" * 140)
        print("✓ Program completed successfully")
        if dry_run:
            print("ℹ️  This was a DRY RUN - no Slack notifications or MariaDB updates were made")
            print("   To send notifications and store data, run without --dry-run flag")
        print("=" * 140)

        return 0

    except Exception as e:
        print(f"\n✗ Error: {e}", file=sys.stderr)
        import traceback
        traceback.print_exc()

        # Close MariaDB session if it exists
        if 'mariadb_session' in locals() and mariadb_session:
            try:
                mariadb_session.close()
            except:
                pass

        return 1


if __name__ == "__main__":
    sys.exit(main())

