#!/usr/bin/env python3
"""
Test query for APR joined data - limited output
"""

import json
import sys
import re
import hashlib
from pathlib import Path
from datetime import datetime, date
import pymssql
import requests
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    print("Warning: openpyxl not installed. Excel export will not be available.")


def load_config() -> dict:
    """Load database configuration from databases_config.json."""
    config_paths = [
        Path.cwd() / "databases_config.json",  # Current working directory (for standalone executables)
        Path(__file__).parent / "databases_config.json",
        Path(__file__).parent.parent / "databases_config.json",
    ]

    for config_path in config_paths:
        if config_path.exists():
            with open(config_path, 'r') as f:
                return json.load(f)

    raise FileNotFoundError(
        f"Configuration file not found in: {[str(p) for p in config_paths]}"
    )


def load_slack_config() -> dict:
    """Load Slack configuration from slack_config.json."""
    config_paths = [
        Path.cwd() / "slack_config.json",  # Current working directory (for standalone executables)
        Path(__file__).parent / "slack_config.json",
        Path(__file__).parent.parent / "slack_config.json",
    ]

    for config_path in config_paths:
        if config_path.exists():
            with open(config_path, 'r') as f:
                return json.load(f)

    # Return default config if file not found
    return {
        "slack": {
            "webhook_url": None,
            "channel": "#alerts",
            "username": "BMA Alert Bot"
        }
    }


def get_mssql_connection():
    """Create connection to MSSQL database."""
    config = load_config()
    cfg = config['databases']['mssql']

    print(f"Connecting to MSSQL server: {cfg['server']}:{cfg['port']}")
    print(f"Database: {cfg['scheme']}")

    conn = pymssql.connect(
        server=cfg['server'],
        port=cfg['port'],
        user=cfg['user'],
        password=cfg['password'],
        database=cfg['scheme'],
        login_timeout=15
    )

    return conn


def validate_duplicata_format(duplicata: str, nfe: int) -> bool:
    """
    Validate DUPLICATA nomenclature.

    Format should be: NFE (with or without leading zeros) + separator (/ or -) + sequential number
    Examples:
        - 021467-1, 021467-2 (valid for NFE=21467)
        - 21467-1, 21467-2 (valid for NFE=21467)
        - 140917/01, 140917/02 (valid for NFE=140917)

    Args:
        duplicata: The DUPLICATA field value
        nfe: The NFE integer value

    Returns:
        True if valid format, False otherwise
    """
    if not duplicata or not nfe:
        return False

    # Pattern: NFE (with optional leading zeros) + separator (/ or -) + sequential number (1+ digits)
    # The NFE part should match the actual NFE value (with or without leading zeros)
    pattern = rf'^0*{nfe}[-/]\d+$'

    return bool(re.match(pattern, duplicata.strip()))


def get_invalid_reason(duplicata: str, nfe: int) -> str:
    """
    Get the reason why DUPLICATA format is invalid (in Brazilian Portuguese).

    Args:
        duplicata: The DUPLICATA field value
        nfe: The NFE integer value

    Returns:
        String describing the error reason in Portuguese
    """
    if not duplicata:
        return "DUPLICATA vazio ou nulo"

    if not nfe:
        return "NFE vazio ou nulo"

    duplicata = duplicata.strip()

    # Check if it contains the NFE number
    if str(nfe) not in duplicata and not re.search(rf'0*{nfe}', duplicata):
        return f"DUPLICATA não contém o número da NFE ({nfe})"

    # Check if it has a separator
    if '-' not in duplicata and '/' not in duplicata:
        return "Falta separador (- ou /) entre NFE e número sequencial"

    # Check if NFE is at the beginning
    if not re.match(rf'^0*{nfe}', duplicata):
        return f"NFE ({nfe}) não está no início da DUPLICATA"

    # Check if there's a sequential number after separator
    if not re.search(r'[-/]\d+$', duplicata):
        return "Falta número sequencial após o separador"

    # Generic error
    return "Formato inválido - esperado: NFE + separador (- ou /) + número sequencial"


def suggest_corrected_duplicata(duplicata: str, nfe: int) -> str:
    """
    Suggest a corrected DUPLICATA format when possible.

    Args:
        duplicata: The DUPLICATA field value
        nfe: The NFE integer value

    Returns:
        Suggested corrected DUPLICATA or empty string if cannot be corrected
    """
    if not duplicata or not nfe:
        return ""

    duplicata = duplicata.strip()
    nfe_str = str(nfe)

    # First, check if DUPLICATA contains the NFE number at all
    # Check both with and without leading zeros
    contains_nfe = False

    # Check if NFE appears in DUPLICATA (with or without leading zeros)
    if nfe_str in duplicata:
        contains_nfe = True
    elif re.search(rf'0*{nfe}', duplicata):
        # Check if it matches with leading zeros at the beginning
        if re.match(rf'^0*{nfe}', duplicata):
            contains_nfe = True

    # If DUPLICATA doesn't contain NFE, we cannot suggest a correction
    if not contains_nfe:
        return ""

    # Case 1: DUPLICATA is just the NFE without separator and sequential
    # Example: "727775" -> "727775-1"
    if duplicata == nfe_str or duplicata == nfe_str.zfill(len(duplicata)):
        return f"{duplicata}-1"

    # Case 2: DUPLICATA has NFE + sequential number concatenated (no separator)
    # Example: "0014687203" (NFE=146872) -> "0014687-03" or "146872-03"
    # Try to find where NFE ends in the duplicata
    if nfe_str in duplicata:
        # Find the position where NFE appears
        nfe_pos = duplicata.find(nfe_str)
        if nfe_pos == 0 or duplicata[:nfe_pos] == '0' * nfe_pos:
            # NFE is at the beginning (with or without leading zeros)
            # Extract the part after NFE as sequential number
            after_nfe = duplicata[nfe_pos + len(nfe_str):]
            if after_nfe and after_nfe.isdigit():
                # Use the original NFE part (with leading zeros if present) + separator + sequential
                nfe_part = duplicata[:nfe_pos + len(nfe_str)]
                return f"{nfe_part}-{after_nfe}"

    # Case 3: Try with leading zeros pattern
    # Example: "0015222401" where NFE=152224
    match = re.match(rf'^(0*{nfe})(\d+)$', duplicata)
    if match:
        nfe_part = match.group(1)
        sequential = match.group(2)
        if sequential:
            return f"{nfe_part}-{sequential}"

    # Case 4: Has separator but wrong position or format
    if '-' in duplicata or '/' in duplicata:
        # Try to extract sequential number
        parts = re.split(r'[-/]', duplicata)
        if len(parts) >= 2:
            # Check if first part contains NFE
            if nfe_str in parts[0] or re.match(rf'^0*{nfe}$', parts[0]):
                # Use the NFE with leading zeros if original had them
                if duplicata.startswith('0') and not nfe_str.startswith('0'):
                    # Preserve leading zeros
                    nfe_with_zeros = duplicata[:duplicata.find('-') if '-' in duplicata else duplicata.find('/')]
                    if nfe_str in nfe_with_zeros or re.match(rf'^0*{nfe}$', nfe_with_zeros):
                        return f"{nfe_with_zeros}-{parts[-1]}"

    # Cannot determine correction
    return ""


def separate_valid_invalid_records(results: list[dict]) -> tuple[list[dict], list[dict]]:
    """
    Separate records into valid and invalid based on DUPLICATA format.

    Args:
        results: List of query results

    Returns:
        Tuple of (valid_records, invalid_records)
    """
    valid_records = []
    invalid_records = []

    for record in results:
        duplicata = str(record.get('DUPLICATA', '')).strip()
        nfe = record.get('NFE', 0)

        if validate_duplicata_format(duplicata, nfe):
            valid_records.append(record)
        else:
            invalid_records.append(record)

    return valid_records, invalid_records


def export_invalid_to_excel(invalid_records: list[dict], filename: str = None) -> str:
    """
    Export invalid records to Excel file.

    Args:
        invalid_records: List of invalid records
        filename: Output filename (optional, will generate if not provided)

    Returns:
        Path to the created Excel file
    """
    if not EXCEL_AVAILABLE:
        print("Error: openpyxl not installed. Cannot export to Excel.")
        return None

    if not filename:
        # Create output directory if it doesn't exist
        output_dir = Path(__file__).parent / 'duplicatas_invalidas'
        output_dir.mkdir(exist_ok=True)

        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = str(output_dir / f'duplicatas_invalidas_{timestamp}.xlsx')

    # Create workbook and worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Duplicatas Inválidas"

    # Define headers
    headers = [
        'GERENTE',
        'PROPOSTA',
        'DATA_OPERACAO',
        'EMPRESA',
        'DUPLICATA',
        'DUPLICATA_CORRETA',
        'VALOR',
        'ID_PRODUTO',
        'NFE',
        'MOTIVO_INVALIDO'
    ]

    # Style for header
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center")

    # Write headers
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    # Get current date in yyyy-MM-dd format
    data_operacao = datetime.now().strftime('%Y-%m-%d')

    # Write data
    for row_num, record in enumerate(invalid_records, 2):
        duplicata = str(record.get('DUPLICATA', '')).strip()
        nfe = record.get('NFE', 0)

        ws.cell(row=row_num, column=1).value = str(record.get('GERENTE', ''))
        ws.cell(row=row_num, column=2).value = record.get('PROPOSTA', '')
        ws.cell(row=row_num, column=3).value = data_operacao
        ws.cell(row=row_num, column=4).value = str(record.get('EMPRESA', ''))
        ws.cell(row=row_num, column=5).value = duplicata
        ws.cell(row=row_num, column=6).value = suggest_corrected_duplicata(duplicata, nfe)
        ws.cell(row=row_num, column=7).value = record.get('VALOR', 0)
        ws.cell(row=row_num, column=8).value = record.get('ID_PRODUTO', '')
        ws.cell(row=row_num, column=9).value = nfe
        ws.cell(row=row_num, column=10).value = get_invalid_reason(duplicata, nfe)

        # Format VALOR as currency
        ws.cell(row=row_num, column=7).number_format = 'R$ #,##0.00'

    # Adjust column widths
    column_widths = {
        'A': 20,  # GERENTE
        'B': 12,  # PROPOSTA
        'C': 16,  # DATA_OPERACAO
        'D': 15,  # EMPRESA
        'E': 18,  # DUPLICATA
        'F': 20,  # DUPLICATA_CORRETA
        'G': 18,  # VALOR
        'H': 14,  # ID_PRODUTO
        'I': 12,  # NFE
        'J': 60   # MOTIVO_INVALIDO
    }

    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # Save workbook
    wb.save(filename)

    return filename


def get_records_hash(records: list[dict]) -> str:
    """
    Generate a hash of the invalid records to track if they've been sent.

    Args:
        records: List of invalid records

    Returns:
        SHA256 hash of the records
    """
    # Create a string representation of key fields from records
    record_keys = []
    for record in records:
        key = f"{record.get('PROPOSTA')}_{record.get('DUPLICATA')}_{record.get('NFE')}"
        record_keys.append(key)

    # Sort to ensure consistent hash
    record_keys.sort()
    records_str = '|'.join(record_keys)

    # Generate hash
    return hashlib.sha256(records_str.encode()).hexdigest()


def get_tracking_file() -> Path:
    """Get the path to the tracking file for today."""
    tracking_dir = Path(__file__).parent / 'duplicatas_invalidas'
    tracking_dir.mkdir(exist_ok=True)

    today = date.today().strftime('%Y-%m-%d')
    return tracking_dir / f'.tracking_{today}.json'


def is_already_sent(records: list[dict]) -> bool:
    """
    Check if these records have already been sent today.

    Args:
        records: List of invalid records

    Returns:
        True if already sent, False otherwise
    """
    tracking_file = get_tracking_file()

    if not tracking_file.exists():
        return False

    try:
        with open(tracking_file, 'r') as f:
            tracking_data = json.load(f)

        current_hash = get_records_hash(records)
        return tracking_data.get('records_hash') == current_hash
    except Exception as e:
        print(f"Warning: Error reading tracking file: {e}")
        return False


def mark_as_sent(records: list[dict]):
    """
    Mark these records as sent in the tracking file.

    Args:
        records: List of invalid records
    """
    tracking_file = get_tracking_file()

    tracking_data = {
        'records_hash': get_records_hash(records),
        'sent_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        'record_count': len(records)
    }

    try:
        with open(tracking_file, 'w') as f:
            json.dump(tracking_data, f, indent=2)
    except Exception as e:
        print(f"Warning: Error writing tracking file: {e}")


def cleanup_old_files():
    """
    Remove Excel files and tracking files that are not from today.
    """
    duplicatas_dir = Path(__file__).parent / 'duplicatas_invalidas'

    if not duplicatas_dir.exists():
        return

    today = date.today().strftime('%Y-%m-%d')
    today_prefix = today.replace('-', '')  # Format: YYYYMMDD

    removed_count = 0

    # Remove old Excel files
    for excel_file in duplicatas_dir.glob('duplicatas_invalidas_*.xlsx'):
        # Extract date from filename (format: duplicatas_invalidas_YYYYMMDD_HHMMSS.xlsx)
        filename = excel_file.name
        if filename.startswith('duplicatas_invalidas_'):
            file_date = filename.split('_')[2]  # Get YYYYMMDD part
            if file_date != today_prefix:
                try:
                    excel_file.unlink()
                    removed_count += 1
                    print(f"  Removed old file: {excel_file.name}")
                except Exception as e:
                    print(f"  Warning: Could not remove {excel_file.name}: {e}")

    # Remove old tracking files
    for tracking_file in duplicatas_dir.glob('.tracking_*.json'):
        # Extract date from filename (format: .tracking_YYYY-MM-DD.json)
        filename = tracking_file.name
        if filename.startswith('.tracking_'):
            file_date = filename.replace('.tracking_', '').replace('.json', '')
            if file_date != today:
                try:
                    tracking_file.unlink()
                    removed_count += 1
                    print(f"  Removed old tracking: {tracking_file.name}")
                except Exception as e:
                    print(f"  Warning: Could not remove {tracking_file.name}: {e}")

    if removed_count > 0:
        print(f"✓ Cleaned up {removed_count} old file(s)")
    else:
        print("✓ No old files to clean up")


def get_channel_id(bot_token: str, channel_name: str) -> str:
    """
    Get channel ID from channel name.

    Args:
        bot_token: Slack bot token
        channel_name: Channel name (with or without #)

    Returns:
        Channel ID or the original channel_name if not found
    """
    try:
        # Remove # if present
        channel_name_clean = channel_name.lstrip('#')

        # Get list of channels
        response = requests.get(
            'https://slack.com/api/conversations.list',
            headers={
                'Authorization': f'Bearer {bot_token}',
                'Content-Type': 'application/json'
            },
            params={
                'types': 'public_channel,private_channel',
                'limit': 1000
            },
            timeout=10
        )

        result = response.json()

        if result.get('ok'):
            for channel in result.get('channels', []):
                if channel.get('name') == channel_name_clean:
                    return channel.get('id')

        # If not found, return original (might be an ID already)
        return channel_name

    except Exception as e:
        print(f"  Warning: Could not resolve channel name: {e}")
        return channel_name


def send_slack_notification(excel_file: str, record_count: int) -> bool:
    """
    Send notification to Slack with Excel file attached.

    Args:
        excel_file: Path to the Excel file
        record_count: Number of invalid records

    Returns:
        True if sent successfully, False otherwise
    """
    try:
        slack_config = load_slack_config()
        bot_token = slack_config['slack'].get('bot_token')
        channel = slack_config['slack'].get('channel', '#alertascredito1')

        if not bot_token or bot_token == "YOUR_SLACK_BOT_TOKEN_HERE":
            print("⚠ Slack bot token not configured. Skipping Slack notification.")
            print("  To enable Slack notifications, update slack_config.json with your bot token.")
            return False

        # Get channel ID
        channel_id = get_channel_id(bot_token, channel)

        # Format current time
        current_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

        # Create initial message text
        message_text = (
            f"*DUPLICATAS com formato inválido*\n"
            f"Quantidade: {record_count}\n"
            f"Data: {current_time}\n"
            f"Relatório: {Path(excel_file).name}"
        )

        # Step 1: Get upload URL
        print(f"  Uploading file to Slack channel: {channel} (ID: {channel_id})")

        filename = Path(excel_file).name
        file_size = Path(excel_file).stat().st_size

        # Get upload URL using form data (not JSON)
        upload_url_response = requests.post(
            'https://slack.com/api/files.getUploadURLExternal',
            headers={
                'Authorization': f'Bearer {bot_token}'
            },
            data={
                'filename': filename,
                'length': file_size
            },
            timeout=30
        )

        upload_url_result = upload_url_response.json()

        if not upload_url_result.get('ok'):
            error_msg = upload_url_result.get('error', 'Unknown error')
            print(f"⚠ Failed to get upload URL: {error_msg}")
            return False

        upload_url = upload_url_result['upload_url']
        file_id = upload_url_result['file_id']

        # Step 2: Upload file to the URL
        with open(excel_file, 'rb') as file_content:
            upload_response = requests.post(
                upload_url,
                data=file_content.read(),
                timeout=60
            )

        if upload_response.status_code != 200:
            print(f"⚠ File upload failed: HTTP {upload_response.status_code}")
            return False

        # Step 3: Complete the upload and share to channel using form data
        complete_response = requests.post(
            'https://slack.com/api/files.completeUploadExternal',
            headers={
                'Authorization': f'Bearer {bot_token}'
            },
            data={
                'files': json.dumps([{'id': file_id, 'title': filename}]),
                'channel_id': channel_id,
                'initial_comment': message_text
            },
            timeout=30
        )

        result = complete_response.json()

        if result.get('ok'):
            print(f"✓ Slack notification sent successfully to {channel}!")
            print(f"  File uploaded: {Path(excel_file).name}")
            return True
        else:
            error_msg = result.get('error', 'Unknown error')
            print(f"⚠ Slack notification failed: {error_msg}")
            if 'needed' in result:
                print(f"  Missing scopes: {result.get('needed')}")
            if 'provided' in result:
                print(f"  Provided scopes: {result.get('provided')}")
            return False

    except FileNotFoundError:
        print(f"⚠ Error: Excel file not found: {excel_file}")
        return False
    except Exception as e:
        print(f"⚠ Error sending Slack notification: {e}")
        import traceback
        traceback.print_exc()
        return False


def main():
    """Main entry point."""
    try:
        print("=" * 130)
        print("APR JOINED DATA - TODAY'S RECORDS (TEST)")
        print("=" * 130)
        
        # Connect to database
        conn = get_mssql_connection()
        print("✓ Connected successfully!\n")
        
        # Query joined data
        query = """
        SELECT
            c.Gerente AS GERENTE,
            t.NUMERO AS PROPOSTA,
            c.empresa AS EMPRESA,
            t.TITULO AS DUPLICATA,
            t.VALOR,
            t.id_produto AS ID_PRODUTO,
            CAST(SUBSTRING(t.NFECHAVE, 26, 9) AS INT) AS NFE
        FROM APR_TITULOS t WITH (NOLOCK)
        INNER JOIN APR_capa c WITH (NOLOCK)
            ON t.NUMERO = c.numero
        WHERE t.DATA >= CAST(GETDATE() AS DATE)
          AND t.DATA < DATEADD(DAY, 1, CAST(GETDATE() AS DATE))
          AND c.Data >= CAST(GETDATE() AS DATE)
          AND c.Data < DATEADD(DAY, 1, CAST(GETDATE() AS DATE))
          AND t.NFECHAVE IS NOT NULL
          AND t.NFECHAVE <> ''
        ORDER BY c.Gerente, t.NUMERO, t.TITULO
        """
        
        cursor = conn.cursor(as_dict=True)
        print("Executing query...")
        cursor.execute(query)
        results = cursor.fetchall()
        cursor.close()

        print(f"Total records: {len(results)}\n")

        # Separate valid and invalid records
        print("Validating DUPLICATA format...")
        valid_records, invalid_records = separate_valid_invalid_records(results)

        print(f"Valid records: {len(valid_records)}")
        print(f"Invalid records: {len(invalid_records)}\n")

        # Print valid records table
        print("=" * 130)
        print("VALID RECORDS (DUPLICATA format matches NFE)")
        print("=" * 130)
        print(f"{'#':<5} {'GERENTE':<20} {'PROPOSTA':<10} {'EMPRESA':<12} {'DUPLICATA':<15} {'VALOR':>15} {'ID_PRODUTO':<12} {'NFE':<10}")
        print("-" * 130)

        # Print first 50 valid records
        for i, row in enumerate(valid_records[:50], 1):
            gerente = str(row.get('GERENTE', ''))[:19]
            proposta = str(row.get('PROPOSTA', ''))
            empresa = str(row.get('EMPRESA', ''))[:11]
            duplicata = str(row.get('DUPLICATA', ''))[:14]
            valor = row.get('VALOR', 0)
            id_produto = str(row.get('ID_PRODUTO', ''))
            nfe = row.get('NFE', 0)

            valor_str = f"R$ {valor:,.2f}" if valor else "R$ 0.00"

            print(f"{i:<5} {gerente:<20} {proposta:<10} {empresa:<12} {duplicata:<15} {valor_str:>15} {id_produto:<12} {nfe:<10}")

        if len(valid_records) > 50:
            print(f"\n... and {len(valid_records) - 50} more valid records")

        print("-" * 130)

        # Calculate total for valid records
        total_valid = sum(row.get('VALOR', 0) for row in valid_records)
        print(f"\n{'VALID TOTAL:':>67} R$ {total_valid:,.2f}")
        print("=" * 130)

        # Print invalid records table
        if invalid_records:
            print("\n")
            print("=" * 130)
            print("INVALID RECORDS (DUPLICATA format does NOT match NFE)")
            print("=" * 130)
            print(f"{'#':<5} {'GERENTE':<20} {'PROPOSTA':<10} {'EMPRESA':<12} {'DUPLICATA':<15} {'VALOR':>15} {'ID_PRODUTO':<12} {'NFE':<10}")
            print("-" * 130)

            # Print all invalid records (or first 100 if too many)
            for i, row in enumerate(invalid_records[:100], 1):
                gerente = str(row.get('GERENTE', ''))[:19]
                proposta = str(row.get('PROPOSTA', ''))
                empresa = str(row.get('EMPRESA', ''))[:11]
                duplicata = str(row.get('DUPLICATA', ''))[:14]
                valor = row.get('VALOR', 0)
                id_produto = str(row.get('ID_PRODUTO', ''))
                nfe = row.get('NFE', 0)

                valor_str = f"R$ {valor:,.2f}" if valor else "R$ 0.00"

                print(f"{i:<5} {gerente:<20} {proposta:<10} {empresa:<12} {duplicata:<15} {valor_str:>15} {id_produto:<12} {nfe:<10}")

            if len(invalid_records) > 100:
                print(f"\n... and {len(invalid_records) - 100} more invalid records")

            print("-" * 130)

            # Calculate total for invalid records
            total_invalid = sum(row.get('VALOR', 0) for row in invalid_records)
            print(f"\n{'INVALID TOTAL:':>67} R$ {total_invalid:,.2f}")
            print("=" * 130)

        # Grand total
        print(f"\n{'GRAND TOTAL:':>67} R$ {total_valid + sum(row.get('VALOR', 0) for row in invalid_records):,.2f}")
        print("=" * 130)

        # Process invalid records
        if invalid_records:
            print("\n")
            print("=" * 130)

            # Cleanup old files first
            print("Cleaning up old files...")
            cleanup_old_files()
            print()

            # Check if already sent today
            if is_already_sent(invalid_records):
                print("⚠ These records have already been sent today. Skipping export and notification.")
                print("=" * 130)
            else:
                # Export to Excel
                print("Exporting invalid records to Excel...")
                excel_file = export_invalid_to_excel(invalid_records)

                if excel_file:
                    print(f"✓ Excel file created: {excel_file}")
                    print(f"  Total invalid records exported: {len(invalid_records)}")
                    print()

                    # Send Slack notification
                    print("Sending Slack notification...")
                    slack_sent = send_slack_notification(excel_file, len(invalid_records))
                    print()

                    # Mark as sent
                    if slack_sent or True:  # Mark as sent even if Slack fails
                        mark_as_sent(invalid_records)
                        print("✓ Records marked as sent for today")
                else:
                    print("✗ Failed to create Excel file")

                print("=" * 130)
        else:
            print("\n")
            print("=" * 130)
            print("✓ No invalid records found!")
            print("=" * 130)

        # Close connection
        conn.close()

        return 0
        
    except Exception as e:
        print(f"\n✗ Error: {e}", file=sys.stderr)
        import traceback
        traceback.print_exc()
        return 1


if __name__ == "__main__":
    sys.exit(main())

