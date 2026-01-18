#!/usr/bin/env python3
"""
Query pre-impresso boleto information joined with APR_TITULOS and APR_capa.
Extracts: GERENTE, PROPOSTA (NUMERO), DATA, CEDENTE, SEUNO, FAIXA
"""

import json
import sys
import re
from pathlib import Path
from datetime import datetime, date, timedelta
from sqlalchemy import create_engine, or_, and_, func, cast, Date
from sqlalchemy.orm import sessionmaker
from models import Cedente, APRTitulos, APRCapa
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import requests
import requests


def get_most_recent_weekday(target_date=None):
    """
    Get the most recent weekday (Monday-Friday).
    If target_date is a weekend, return the most recent Friday.
    """
    if target_date is None:
        target_date = date.today()

    # weekday(): Monday=0, Sunday=6
    # Saturday=5, Sunday=6
    while target_date.weekday() >= 5:  # Saturday or Sunday
        target_date -= timedelta(days=1)

    return target_date


def calculate_verification_digit(seuno):
    """
    Calculate the verification digit for a SEUNO.

    Algorithm:
    1. Form seuno_vector: [1, 9] + first 11 digits of SEUNO
    2. factor_vector: [2, 7, 6, 5, 4, 3, 2, 7, 6, 5, 4, 3, 2]
    3. dot_product = sum(seuno_vector[i] * factor_vector[i])
    4. quotient = dot_product // 11
    5. pre_last_remainder = dot_product - (quotient * 11)  [this is the modulo/remainder]
    6. Calculate verification_digit:
       - if (11 - pre_last_remainder) == 11: verification_digit = 0
       - if (11 - pre_last_remainder) == 10: verification_digit = 'P'
       - else: verification_digit = 11 - pre_last_remainder

    Example:
        SEUNO = '518300121527'
        seuno_vector = [1, 9, 5, 1, 8, 3, 0, 0, 1, 2, 1, 5, 2]
        factor_vector = [2, 7, 6, 5, 4, 3, 2, 7, 6, 5, 4, 3, 2]
        dot_product = 180
        quotient = 180 // 11 = 16
        pre_last_remainder = 180 - (16 * 11) = 180 - 176 = 4
        verification_digit = 11 - 4 = 7

    Args:
        seuno: string representing the SEUNO (e.g., '518300121527')

    Returns:
        verification_digit as string (digit or 'P')
    """
    if not seuno or len(seuno) < 11:
        return ''

    # Extract first 11 digits from SEUNO (ignore non-digit characters)
    seuno_digits = [int(c) for c in seuno if c.isdigit()][:11]

    if len(seuno_digits) < 11:
        return ''

    # Form seuno_vector: [1, 9] + first 11 digits
    seuno_vector = [1, 9] + seuno_digits

    # Fixed factor_vector
    factor_vector = [2, 7, 6, 5, 4, 3, 2, 7, 6, 5, 4, 3, 2]

    # Calculate dot product
    dot_product = sum(s * f for s, f in zip(seuno_vector, factor_vector))

    # Calculate quotient (integer division)
    quotient = dot_product // 11

    # Calculate pre_last_remainder (this is the modulo: dot_product % 11)
    pre_last_remainder = dot_product - (quotient * 11)

    # Calculate verification_digit based on rules
    result = 11 - pre_last_remainder

    if result == 11:
        verification_digit = '0'
    elif result == 10:
        verification_digit = 'P'
    else:
        verification_digit = str(result)

    return verification_digit


def validate_seuno(seuno, range_value, calculated_verif):
    """
    Validate SEUNO according to business rules.

    Rules:
    1. SEUNO must start with the RANGE value
    2. After the first 11 digits, it must contain a valid verification digit (0, P, or 2-9)

    Args:
        seuno: string representing the SEUNO (e.g., '518300121527')
        range_value: string representing the expected range (e.g., '5183')
        calculated_verif: string representing the calculated verification digit

    Returns:
        tuple (is_valid: bool, motivo_invalido: str)
    """
    if not seuno:
        return False, "SEUNO vazio ou nulo"

    # Remove non-digit characters for validation (except P)
    seuno_clean = ''.join(c for c in seuno if c.isdigit() or c.upper() == 'P')

    if len(seuno_clean) < 12:
        return False, "SEUNO com menos de 12 caracteres"

    # Rule 1: SEUNO must start with RANGE value
    if not seuno_clean.startswith(range_value):
        return False, f"SEUNO não inicia com a faixa {range_value}"

    # Rule 2: Position 12 (index 11) must contain a valid verification digit
    actual_verif = seuno_clean[11] if len(seuno_clean) > 11 else ''

    # Valid verification digits: 0, P, 1-9
    valid_digits = {'0', 'P', '1', '2', '3', '4', '5', '6', '7', '8', '9'}

    if actual_verif not in valid_digits:
        return False, f"Dígito verificador inválido na posição 12: '{actual_verif}'"

    # Check if the actual verification digit matches the calculated one
    if actual_verif != calculated_verif:
        return False, f"Dígito verificador incorreto: esperado '{calculated_verif}', encontrado '{actual_verif}'"

    return True, ""


def load_config() -> dict:
    """Load database configuration."""
    config_paths = [
        Path.cwd() / "databases_config.json",  # Current working directory (for standalone executables)
        Path(__file__).parent / "databases_config.json",
        Path(__file__).parent.parent / "databases_config.json",
    ]

    for config_path in config_paths:
        if config_path.exists():
            with open(config_path, 'r') as f:
                return json.load(f)

    raise FileNotFoundError("Configuration file not found")


def load_slack_config() -> dict:
    """Load Slack configuration from slack_config.json."""
    config_paths = [
        Path.cwd() / "slack_config.json",  # Current working directory (for standalone executables)
        Path(__file__).parent / "slack_config.json",
        Path(__file__).parent.parent / "slack_config.json",
    ]

    for config_path in config_paths:
        if config_path.exists():
            with open(config_path, 'r') as f:
                return json.load(f)

    # Return default config if file not found
    return {
        "slack": {
            "bot_token": None,
            "channel": "#alerts",
            "username": "BMA Alert Bot"
        }
    }


def create_db_engine():
    """Create SQLAlchemy engine for MSSQL database."""
    config = load_config()
    cfg = config['databases']['mssql']
    
    connection_string = (
        f"mssql+pymssql://{cfg['user']}:{cfg['password']}"
        f"@{cfg['server']}:{cfg['port']}/{cfg['scheme']}"
    )
    
    engine = create_engine(connection_string, echo=False)
    return engine


def extract_company_range_pairs(text: str) -> list[dict]:
    """
    Extract ALL company/range pairs from pre-impresso text.
    
    Returns list of dictionaries with: company, range
    """
    if not text:
        return []
    
    results = []
    
    # Company patterns (case insensitive)
    company_patterns = [
        (r'BMA\s*FIDC', 'BMA FIDC'),
        (r'BMA\s*INTER', 'BMA INTER'),
        (r'BMA\s*SEC', 'BMA SEC'),
        (r'BMA\s*CAPITAL', 'BMA CAPITAL'),
        (r'Fidc', 'BMA FIDC'),
        (r'Inter(?!\s*(?:company|nacional))', 'BMA INTER'),
    ]
    
    # Also look for standalone "BMA"
    company_patterns.append((r'BMA(?!\s*(?:FIDC|INTER|SEC|CAPITAL|Fidc|Inter))', 'BMA'))
    
    for pattern, normalized_name in company_patterns:
        for match in re.finditer(pattern, text, re.IGNORECASE):
            company_start = match.start()
            company_end = match.end()
            
            lookahead_text = text[company_end:company_end + 120]
            
            # Try multiple patterns to find the range number
            faixa_pattern = r'(?:faixa|n/n|numero|nº|n°)[\s:\-]*(\d{1,5}|esquerda)'
            faixa_match = re.search(faixa_pattern, lookahead_text, re.IGNORECASE)
            
            if faixa_match:
                range_value = faixa_match.group(1)
                results.append({
                    'company': normalized_name,
                    'range': range_value,
                    'position': company_start
                })
                continue
            
            # Direct number after separators
            direct_pattern = r'[\s\-:]+(\d{2,5})(?:\s|$|\.|\||/)'
            direct_match = re.search(direct_pattern, lookahead_text)
            
            if direct_match:
                range_value = direct_match.group(1)
                results.append({
                    'company': normalized_name,
                    'range': range_value,
                    'position': company_start
                })
                continue
            
            # Number after "N/N"
            nn_pattern = r'N/N[\s:\-]*(\d{1,5})'
            nn_match = re.search(nn_pattern, lookahead_text, re.IGNORECASE)
            
            if nn_match:
                range_value = nn_match.group(1)
                results.append({
                    'company': normalized_name,
                    'range': range_value,
                    'position': company_start
                })
                continue
    
    # Sort by position and remove duplicates
    results.sort(key=lambda x: x['position'])
    
    seen = set()
    unique_results = []
    for item in results:
        key = (item['company'], item['range'])
        if key not in seen:
            seen.add(key)
            unique_results.append({
                'company': item['company'],
                'range': item['range']
            })
    
    return unique_results


def normalize_empresa_to_company(empresa):
    """
    Map the empresa field from APR_capa to the company names used in pre-impresso.

    Args:
        empresa: string from APR_capa.empresa field (e.g., 'FIDC', 'BMA', 'SEC', 'INTER')

    Returns:
        Normalized company name (e.g., 'BMA FIDC', 'BMA', 'BMA SEC', 'BMA INTER')
    """
    if not empresa:
        return None

    empresa_upper = empresa.strip().upper()

    # Map empresa values to company names
    empresa_mapping = {
        'FIDC': 'BMA FIDC',
        'BMA': 'BMA',
        'SEC': 'BMA SEC',
        'INTER': 'BMA INTER',
        'BMA FIDC': 'BMA FIDC',
        'BMA SEC': 'BMA SEC',
        'BMA INTER': 'BMA INTER',
    }

    return empresa_mapping.get(empresa_upper, None)


def query_pre_impresso_with_propostas(session, target_date=None):
    """
    Query cedentes with pre-impresso, join with APR_TITULOS and APR_capa.

    Returns: GERENTE, PROPOSTA, DATA, CEDENTE, SEUNO, RANGE
    """
    if target_date is None:
        target_date = date.today()

    # Ensure we use a weekday
    original_date = target_date
    target_date = get_most_recent_weekday(target_date)

    if original_date != target_date:
        print(f"\n⚠️  {original_date} is a weekend. Using most recent weekday: {target_date}")

    print("=" * 140)
    print(f"PRE-IMPRESSO BOLETOS WITH PROPOSTAS - DATE: {target_date} ({target_date.strftime('%A')})")
    print("=" * 140)

    # Step 1: Get cedentes with pre-impresso and their ranges
    print("\nStep 1: Extracting cedentes with pre-impresso...")
    cedentes_with_pre_impresso = session.query(Cedente).filter(
        Cedente.obs1 != None,
        or_(
            Cedente.obs1.like('%pre%impresso%'),
            Cedente.obs1.like('%pré%impresso%'),
            Cedente.obs1.like('%pre%ìmpresso%'),
            Cedente.obs1.like('%pré%ìmpresso%'),
        )
    ).all()

    # Build a mapping: cedente apelido -> list of (company, range) pairs
    cedente_ranges = {}
    for cedente in cedentes_with_pre_impresso:
        obs1_text = str(cedente.obs1) if cedente.obs1 else ""
        pairs = extract_company_range_pairs(obs1_text)
        if pairs:
            cedente_ranges[cedente.apelido] = pairs

    print(f"Found {len(cedente_ranges)} cedentes with pre-impresso ranges")

    # Step 2: Query APR_capa and APR_TITULOS for today's propostas
    print(f"\nStep 2: Querying propostas for date {target_date}...")

    # Join APR_capa with APR_TITULOS on DATA and NUMERO
    # Filter by target date
    # Filter out records with empty/null NFECHAVE
    # Include empresa field to match with company
    # Note: We need to compare only the date part of the datetime column
    query = session.query(
        APRCapa.GERENTE,
        APRCapa.NUMERO.label('PROPOSTA'),
        APRCapa.DATA,
        APRCapa.CEDENTE,
        APRCapa.empresa,
        APRTitulos.SEUNO,
        APRTitulos.TITULO.label('DUPLICATA'),
        APRTitulos.NFEChave,
        Cedente.apelido
    ).join(
        APRTitulos,
        and_(
            APRCapa.DATA == APRTitulos.DATA,
            APRCapa.NUMERO == APRTitulos.NUMERO
        )
    ).join(
        Cedente,
        APRCapa.CEDENTE == Cedente.apelido
    ).filter(
        cast(APRCapa.DATA, Date) == target_date,
        APRCapa.CEDENTE.in_(list(cedente_ranges.keys())),
        APRTitulos.NFEChave.isnot(None),
        APRTitulos.NFEChave != ''
    ).distinct()

    results = query.all()

    print(f"Found {len(results)} proposta records")

    # Step 3: Match empresa with company/range pairs
    print("\nStep 3: Matching empresa with company/range pairs...")

    expanded_results = []
    for row in results:
        gerente = row.GERENTE
        proposta = row.PROPOSTA
        data = row.DATA
        cedente = row.CEDENTE
        empresa = row.empresa
        seuno = row.SEUNO
        duplicata = row.DUPLICATA
        apelido = row.apelido

        # Normalize empresa to company name
        company_from_empresa = normalize_empresa_to_company(empresa)

        # Get all company/range pairs for this cedente
        pairs = cedente_ranges.get(apelido, [])

        # Find the matching pair for this empresa
        matching_pair = None
        for pair in pairs:
            if pair['company'] == company_from_empresa:
                matching_pair = pair
                break

        # If no matching pair found, skip this record or mark as invalid
        if not matching_pair:
            # Create a record with no range validation (empresa doesn't match any pre-impresso company)
            verification_digit = calculate_verification_digit(seuno or '')
            expanded_results.append({
                'GERENTE': gerente,
                'PROPOSTA': proposta,
                'DATA': data,
                'CEDENTE': cedente,
                'SEUNO': seuno,
                'DUPLICATA': duplicata,
                'COMPANY': company_from_empresa or 'DESCONHECIDO',
                'RANGE': '',
                'VERIFICATION_DIGIT': verification_digit,
                'IS_VALID': False,
                'MOTIVO_INVALIDO': f'Empresa {empresa} não possui faixa pré-impressa cadastrada'
            })
            continue

        # Calculate verification digit for this SEUNO
        verification_digit = calculate_verification_digit(seuno or '')

        # Validate SEUNO against the matching range
        is_valid, motivo_invalido = validate_seuno(seuno, matching_pair['range'], verification_digit)

        expanded_results.append({
            'GERENTE': gerente,
            'PROPOSTA': proposta,
            'DATA': data,
            'CEDENTE': cedente,
            'SEUNO': seuno,
            'DUPLICATA': duplicata,
            'COMPANY': matching_pair['company'],
            'RANGE': matching_pair['range'],
            'VERIFICATION_DIGIT': verification_digit,
            'IS_VALID': is_valid,
            'MOTIVO_INVALIDO': motivo_invalido
        })

    # Separate valid and invalid records
    valid_results = [r for r in expanded_results if r['IS_VALID']]
    invalid_results = [r for r in expanded_results if not r['IS_VALID']]

    # Display results
    print(f"\n{'=' * 200}")
    print(f"RESULTS: {len(expanded_results)} records ({len(valid_results)} válidos, {len(invalid_results)} inválidos)")
    print(f"{'=' * 200}\n")

    # Table header
    print(f"{'GERENTE':<12} | {'PROPOSTA':<10} | {'DATA':<12} | {'CEDENTE':<20} | {'SEUNO':<12} | {'DUPLICATA':<14} | {'COMPANY':<15} | {'RANGE':<10} | {'VERIF':<5} | {'MOTIVO_INVALIDO':<60}")
    print("=" * 220)

    # Display all results
    for result in expanded_results:
        gerente = result['GERENTE'] or ''
        proposta = str(result['PROPOSTA'])
        data_str = result['DATA'].strftime('%Y-%m-%d') if result['DATA'] else ''
        cedente = result['CEDENTE'] or ''
        seuno = result['SEUNO'] or ''
        duplicata = result['DUPLICATA'] or ''
        company = result['COMPANY']
        range_val = result['RANGE']
        verif = result['VERIFICATION_DIGIT']
        motivo = result['MOTIVO_INVALIDO'] or ''

        print(f"{gerente:<12} | {proposta:<10} | {data_str:<12} | {cedente:<20} | {seuno:<12} | {duplicata:<14} | {company:<15} | {range_val:<10} | {verif:<5} | {motivo:<60}")

    print("=" * 220)

    # Statistics
    print(f"\n{'=' * 140}")
    print("STATISTICS")
    print(f"{'=' * 140}")
    print(f"Total records: {len(expanded_results)}")
    print(f"Valid records: {len(valid_results)} ({len(valid_results)/len(expanded_results)*100:.1f}%)")
    print(f"Invalid records: {len(invalid_results)} ({len(invalid_results)/len(expanded_results)*100:.1f}%)")
    print(f"Unique propostas: {len(set(r['PROPOSTA'] for r in expanded_results))}")
    print(f"Unique cedentes: {len(set(r['CEDENTE'] for r in expanded_results))}")
    print(f"Unique gerentes: {len(set(r['GERENTE'] for r in expanded_results if r['GERENTE']))}")

    # By company
    company_counts = {}
    for r in expanded_results:
        company = r['COMPANY']
        company_counts[company] = company_counts.get(company, 0) + 1

    print(f"\nBy Company:")
    for company, count in sorted(company_counts.items(), key=lambda x: x[1], reverse=True):
        print(f"  {company}: {count}")

    # Invalid records summary
    if invalid_results:
        print(f"\n{'=' * 140}")
        print(f"INVALID RECORDS SUMMARY ({len(invalid_results)} records)")
        print(f"{'=' * 140}")

        # Group by motivo
        motivo_counts = {}
        for r in invalid_results:
            motivo = r['MOTIVO_INVALIDO']
            motivo_counts[motivo] = motivo_counts.get(motivo, 0) + 1

        for motivo, count in sorted(motivo_counts.items(), key=lambda x: x[1], reverse=True):
            print(f"  {motivo}: {count} records")

    print("=" * 140)

    return expanded_results


def export_invalid_seuno_to_excel(invalid_results, target_date):
    """
    Export invalid SEUNO records to Excel file.

    Args:
        invalid_results: List of dictionaries with invalid SEUNO records
        target_date: Date object for the query date

    Returns:
        Path to the generated Excel file
    """
    if not invalid_results:
        print("\n✓ No invalid records to export!")
        return None

    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "SEUNO Inválidos"

    # Define headers (using EMPRESA and FAIXA instead of COMPANY and RANGE)
    headers = ['GERENTE', 'PROPOSTA', 'DATA', 'CEDENTE', 'SEUNO', 'DUPLICATA', 'EMPRESA', 'FAIXA', 'VERIF', 'MOTIVO_INVALIDO']

    # Style for header
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center")

    # Write headers
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    # Write data
    for row_num, record in enumerate(invalid_results, 2):
        ws.cell(row=row_num, column=1, value=record['GERENTE'] or '')
        ws.cell(row=row_num, column=2, value=record['PROPOSTA'])
        ws.cell(row=row_num, column=3, value=record['DATA'].strftime('%Y-%m-%d') if record['DATA'] else '')
        ws.cell(row=row_num, column=4, value=record['CEDENTE'] or '')
        ws.cell(row=row_num, column=5, value=record['SEUNO'] or '')
        ws.cell(row=row_num, column=6, value=record['DUPLICATA'] or '')
        ws.cell(row=row_num, column=7, value=record['COMPANY'])  # Maps to EMPRESA column
        ws.cell(row=row_num, column=8, value=record['RANGE'])    # Maps to FAIXA column
        ws.cell(row=row_num, column=9, value=record['VERIFICATION_DIGIT'])
        ws.cell(row=row_num, column=10, value=record['MOTIVO_INVALIDO'])

    # Adjust column widths
    column_widths = {
        'A': 15,  # GERENTE
        'B': 12,  # PROPOSTA
        'C': 12,  # DATA
        'D': 25,  # CEDENTE
        'E': 15,  # SEUNO
        'F': 15,  # DUPLICATA
        'G': 18,  # EMPRESA
        'H': 10,  # FAIXA
        'I': 8,   # VERIF
        'J': 60,  # MOTIVO_INVALIDO
    }

    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # Freeze header row
    ws.freeze_panes = 'A2'

    # Create output directory if it doesn't exist
    output_dir = Path(__file__).parent / 'seuno_invalidos'
    output_dir.mkdir(exist_ok=True)

    # Generate filename with timestamp
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    date_str = target_date.strftime('%Y-%m-%d')
    filename = f"seuno_invalidos_{date_str}_{timestamp}.xlsx"
    filepath = output_dir / filename

    # Save workbook
    wb.save(filepath)

    print(f"\n✓ Excel file generated: {filepath}")
    print(f"  - Total invalid records: {len(invalid_results)}")

    return filepath


def cleanup_old_files(current_report_date: date):
    """
    Remove Excel files that are not from the current report date.

    Args:
        current_report_date: The date of the current report being processed
    """
    seuno_dir = Path(__file__).parent / 'seuno_invalidos'

    if not seuno_dir.exists():
        return

    current_date_str = current_report_date.strftime('%Y-%m-%d')

    removed_count = 0

    # Remove old Excel files
    for excel_file in seuno_dir.glob('seuno_invalidos_*.xlsx'):
        # Extract date from filename (format: seuno_invalidos_YYYY-MM-DD_HHMMSS.xlsx)
        filename = excel_file.name
        if filename.startswith('seuno_invalidos_'):
            # Get date part (YYYY-MM-DD)
            parts = filename.replace('seuno_invalidos_', '').split('_')
            if len(parts) >= 1:
                file_date = parts[0]  # YYYY-MM-DD
                if file_date != current_date_str:
                    try:
                        excel_file.unlink()
                        removed_count += 1
                        print(f"  Removed old file: {excel_file.name}")
                    except Exception as e:
                        print(f"  Warning: Could not remove {excel_file.name}: {e}")

    if removed_count > 0:
        print(f"\n✓ Cleaned up {removed_count} old file(s)")
    else:
        print("\n✓ No old files to clean up")


def get_channel_id(bot_token: str, channel_name: str) -> str:
    """
    Get channel ID from channel name.

    Args:
        bot_token: Slack bot token
        channel_name: Channel name (with or without #)

    Returns:
        Channel ID or the original channel_name if not found
    """
    # Remove # if present
    channel_name = channel_name.lstrip('#')

    try:
        response = requests.get(
            'https://slack.com/api/conversations.list',
            headers={'Authorization': f'Bearer {bot_token}'},
            params={'types': 'public_channel,private_channel'},
            timeout=30
        )

        result = response.json()

        if result.get('ok'):
            for channel in result.get('channels', []):
                if channel.get('name') == channel_name:
                    return channel.get('id')

        # If not found, return the original name
        return channel_name

    except Exception as e:
        print(f"  Warning: Could not get channel ID: {e}")
        return channel_name


def send_slack_notification(excel_file: Path, record_count: int, target_date: date) -> bool:
    """
    Send notification to Slack with Excel file attached.

    Args:
        excel_file: Path to the Excel file
        record_count: Number of invalid records
        target_date: Date of the report

    Returns:
        True if sent successfully, False otherwise
    """
    try:
        slack_config = load_slack_config()
        bot_token = slack_config['slack'].get('bot_token')
        channel = slack_config['slack'].get('channel', '#alerts')

        if not bot_token:
            print("\n⚠ Slack bot token not configured. Skipping Slack notification.")
            return False

        print(f"\n📤 Sending Slack notification...")

        # Get channel ID if needed
        channel_id = channel
        if channel.startswith('#') or not channel.startswith('C'):
            channel_id = get_channel_id(bot_token, channel)

        # Create message text
        current_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        date_str = target_date.strftime('%Y-%m-%d')

        message_text = (
            f"*SEUNO Inválidos - Relatório de Pré-Impresso*\n"
            f"📊 Quantidade: {record_count} registros inválidos\n"
            f"📅 Data da operação: {date_str}\n"
            f"🕐 Gerado em: {current_time}\n"
            f"📄 Arquivo: {excel_file.name}"
        )

        # Step 1: Get upload URL
        print(f"  Uploading file to Slack channel: {channel} (ID: {channel_id})")

        filename = excel_file.name
        file_size = excel_file.stat().st_size

        # Get upload URL using form data
        upload_url_response = requests.post(
            'https://slack.com/api/files.getUploadURLExternal',
            headers={'Authorization': f'Bearer {bot_token}'},
            data={
                'filename': filename,
                'length': file_size
            },
            timeout=30
        )

        upload_url_result = upload_url_response.json()

        if not upload_url_result.get('ok'):
            error_msg = upload_url_result.get('error', 'Unknown error')
            print(f"⚠ Failed to get upload URL: {error_msg}")
            return False

        upload_url = upload_url_result['upload_url']
        file_id = upload_url_result['file_id']

        # Step 2: Upload file to the URL
        with open(excel_file, 'rb') as file_content:
            upload_response = requests.post(
                upload_url,
                data=file_content.read(),
                timeout=60
            )

        if upload_response.status_code != 200:
            print(f"⚠ File upload failed: HTTP {upload_response.status_code}")
            return False

        # Step 3: Complete the upload and share to channel
        complete_response = requests.post(
            'https://slack.com/api/files.completeUploadExternal',
            headers={'Authorization': f'Bearer {bot_token}'},
            data={
                'files': json.dumps([{'id': file_id, 'title': filename}]),
                'channel_id': channel_id,
                'initial_comment': message_text
            },
            timeout=30
        )

        result = complete_response.json()

        if result.get('ok'):
            print(f"✓ Slack notification sent successfully to {channel}!")
            print(f"  File uploaded: {excel_file.name}")
            return True
        else:
            error_msg = result.get('error', 'Unknown error')
            print(f"⚠ Slack notification failed: {error_msg}")
            return False

    except Exception as e:
        print(f"\n⚠ Error sending Slack notification: {e}")
        import traceback
        traceback.print_exc()
        return False


def main():
    """Main entry point."""
    try:
        # Check if date argument provided
        target_date = None
        if len(sys.argv) > 1:
            try:
                target_date = datetime.strptime(sys.argv[1], '%Y-%m-%d').date()
                print(f"\nUsing provided date: {target_date}")
            except ValueError:
                print(f"Invalid date format. Use YYYY-MM-DD. Using today's date.")
                target_date = date.today()
        else:
            target_date = date.today()

        print("\nConnecting to database...")
        engine = create_db_engine()
        Session = sessionmaker(bind=engine)
        session = Session()
        print("✓ Connected successfully!\n")

        # Get the most recent weekday
        weekday_date = get_most_recent_weekday(target_date)

        # Query and display results
        results = query_pre_impresso_with_propostas(session, target_date)

        # Filter invalid results
        invalid_results = [r for r in results if not r['IS_VALID']]

        # Export invalid records to Excel
        if invalid_results:
            excel_file = export_invalid_seuno_to_excel(invalid_results, weekday_date)

            # Send Slack notification
            send_slack_notification(excel_file, len(invalid_results), weekday_date)

            # Cleanup old files (keep only files from the current report date)
            cleanup_old_files(weekday_date)
        else:
            print("\n✓ No invalid SEUNO records found!")

        # Close session
        session.close()

        return 0

    except Exception as e:
        print(f"\n✗ Error: {e}", file=sys.stderr)
        import traceback
        traceback.print_exc()
        return 1


if __name__ == "__main__":
    sys.exit(main())


