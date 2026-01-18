#!/usr/bin/env python3
"""
Verify Excel file contents
"""

import openpyxl
import sys

def verify_excel(filename):
    """Verify and display sample of Excel file."""
    wb = openpyxl.load_workbook(filename)
    ws = wb.active
    
    print(f"Excel File: {filename}")
    print(f"Sheet Name: {ws.title}")
    print(f"Total Rows: {ws.max_row - 1} (excluding header)")
    print(f"Total Columns: {ws.max_column}")
    print("\n" + "=" * 150)
    print("SAMPLE DATA (First 10 rows)")
    print("=" * 150)
    
    # Print header
    headers = []
    for col in range(1, ws.max_column + 1):
        headers.append(ws.cell(1, col).value)
    
    print(" | ".join(f"{h:^20}" for h in headers))
    print("-" * 150)
    
    # Print first 10 data rows
    for row in range(2, min(12, ws.max_row + 1)):
        values = []
        for col in range(1, ws.max_column + 1):
            cell_value = ws.cell(row, col).value
            if col == 7:  # VALOR column (now column 7)
                values.append(f"R$ {cell_value:,.2f}" if cell_value else "R$ 0.00")
            else:
                values.append(str(cell_value) if cell_value is not None else "")

        # Truncate long values for display
        display_values = []
        for i, v in enumerate(values):
            if i == 9:  # MOTIVO_INVALIDO column - allow longer
                display_values.append(f"{v[:50]:^20}" if len(v) > 50 else f"{v:^20}")
            else:
                display_values.append(f"{v[:18]:^20}")

        print(" | ".join(display_values))
    
    print("=" * 150)
    
    # Show unique error reasons
    print("\nUNIQUE ERROR REASONS (MOTIVO_INVALIDO):")
    print("-" * 150)
    error_reasons = {}
    for row in range(2, ws.max_row + 1):
        reason = ws.cell(row, 10).value  # MOTIVO_INVALIDO column (now column 10)
        if reason:
            error_reasons[reason] = error_reasons.get(reason, 0) + 1

    for reason, count in sorted(error_reasons.items(), key=lambda x: x[1], reverse=True):
        print(f"  [{count:4d} records] {reason}")

    print("=" * 150)

    # Show correction statistics
    print("\nCORRECTION STATISTICS:")
    print("-" * 150)
    corrected_count = 0
    not_corrected_count = 0
    for row in range(2, ws.max_row + 1):
        corrected = ws.cell(row, 6).value  # DUPLICATA_CORRETA column
        if corrected:
            corrected_count += 1
        else:
            not_corrected_count += 1

    print(f"  Records with suggested correction: {corrected_count}")
    print(f"  Records without correction: {not_corrected_count}")
    print("=" * 150)

if __name__ == "__main__":
    if len(sys.argv) > 1:
        filename = sys.argv[1]
    else:
        # Find the most recent file
        import glob
        files = glob.glob('/tmp/duplicatas_invalidas_*.xlsx')
        if files:
            filename = max(files, key=lambda x: x)
        else:
            print("No Excel file found!")
            sys.exit(1)
    
    verify_excel(filename)

