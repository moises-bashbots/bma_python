#!/usr/bin/env python3
"""
Repurchase Report - Read and process repurchase data from Excel file.

This program reads the "Planilha Cobranca.xlsx" file and processes
repurchase information for clients.

File location: /home/robot/PBI/Cobranca/Planilha Cobranca.xlsx
Main sheet: Sheet1

Columns:
- Gerente (Manager)
- Operador (Operator)
- Grupo (Group)
- Cedente (Client)
- Particularidade (Particularity/Notes)
- Tipo de Recompra (Repurchase Type)
- Prazo de recompra (Repurchase deadline)
- Prazo de recompra definido pelo Rating (Rating-defined deadline)
- Aumentar prazo de recompra para quantos dias? (Increase deadline to how many days?)
- Contato do Cedente (Client contact)
- Restrição em mudar de operador? Sim ou Não (Restriction on changing operator?)
- OBSERVAÇÃO DO CLIENTE (Client observation)
- Parecer resumido do cliente (Client summary opinion)
- Rating 03/09/2024
- FÉRIAS COLETIVAS (Collective vacation)
"""

import sys
from pathlib import Path
from datetime import datetime, date, time, timedelta
from typing import List, Dict, Any, Optional
import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
import unicodedata
import json
import requests
from sqlalchemy import create_engine, cast, Date
from sqlalchemy.orm import sessionmaker
from models import APRCapa, Cedente


# File configuration
EXCEL_FILE_PATH = "/home/robot/PBI/Cobranca/Planilha Cobranca.xlsx"
MAIN_SHEET_NAME = "Sheet1"
DATABASE_CONFIG_FILE = "databases_config.json"
SLACK_CONFIG_FILE = "slack_config.json"
SLACK_CHANNEL_NAME = "operacionalxcobranca"  # Channel name
SENT_ALERTS_FILE = "sent_alerts.json"  # File to track sent alerts


def load_database_config(config_file: str = DATABASE_CONFIG_FILE) -> dict:
    """
    Load database configuration from JSON file.

    Args:
        config_file: Path to the configuration file

    Returns:
        Database configuration dictionary

    Raises:
        FileNotFoundError: If config file doesn't exist
    """
    # Use current working directory for config files (for standalone binary)
    config_path = Path.cwd() / config_file

    if not config_path.exists():
        raise FileNotFoundError(f"Database configuration file not found: {config_path}")

    with open(config_path, 'r') as f:
        config = json.load(f)

    return config['databases']['mssql']


def create_db_engine():
    """
    Create SQLAlchemy engine for MSSQL database.

    Returns:
        SQLAlchemy engine
    """
    config = load_database_config()

    connection_string = (
        f"mssql+pymssql://{config['user']}:{config['password']}"
        f"@{config['server']}:{config['port']}/{config['scheme']}"
    )

    engine = create_engine(connection_string, echo=False)
    return engine


def create_mariadb_connection():
    """
    Create MariaDB connection for dias_uteis queries.

    Returns:
        pymysql connection
    """
    import pymysql

    # Load config from current directory (for standalone binary)
    config_path = Path.cwd() / DATABASE_CONFIG_FILE

    with open(config_path, 'r') as f:
        config = json.load(f)

    db_config = config['databases']['mariadb']

    return pymysql.connect(
        host=db_config['server'],
        port=db_config['port'],
        user=db_config['user'],
        password=db_config['password'],
        database=db_config['scheme'],
        charset='utf8mb4',
        cursorclass=pymysql.cursors.DictCursor
    )


def get_previous_business_day(from_date: date = None) -> Optional[date]:
    """
    Get the previous business day from dias_uteis table.

    Args:
        from_date: Starting date (default: today)

    Returns:
        Previous business day, or None if not found
    """
    if from_date is None:
        from_date = date.today()

    conn = create_mariadb_connection()
    cursor = conn.cursor()

    try:
        cursor.execute('''
            SELECT data FROM dias_uteis
            WHERE data < %s AND eh_dia_util = TRUE
            ORDER BY data DESC
            LIMIT 1
        ''', (from_date,))

        result = cursor.fetchone()
        return result['data'] if result else None
    finally:
        conn.close()


def is_business_day(target_date: date) -> bool:
    """
    Check if a date is a business day using dias_uteis table.

    Args:
        target_date: Date to check

    Returns:
        True if business day, False otherwise
    """
    conn = create_mariadb_connection()
    cursor = conn.cursor()

    try:
        cursor.execute(
            'SELECT eh_dia_util FROM dias_uteis WHERE data = %s',
            (target_date,)
        )
        result = cursor.fetchone()

        if result:
            return bool(result['eh_dia_util'])
        else:
            # If date not in table, assume it's a business day (fallback)
            return target_date.weekday() < 5
    finally:
        conn.close()


def should_exclude_product(produtos: Optional[str]) -> bool:
    """
    Check if a borderô should be excluded from alerts based on its products.

    Excluded products (case-insensitive, accent-insensitive):
    - Capital de Giro
    - Capital de Giro NP
    - Cobrança Simples
    - Cobrança Simples Garantia
    - Nota Comercial
    - CCB
    - Renegociação

    Args:
        produtos: Pipe-separated product names (e.g., "CONVENCIONAL | CHEQUE")

    Returns:
        True if any product in the list should be excluded, False otherwise
    """
    if not produtos:
        return False

    # List of excluded products (normalized: lowercase, no accents)
    EXCLUDED_PRODUCTS = {
        'capital de giro',
        'capital de giro np',
        'cobranca simples',
        'cobranca simples garantia',
        'cob simples garantia',  # Alternative name
        'nota comercial',
        'ccb',
        'renegociacao',
    }

    # Split products by pipe separator
    product_list = [p.strip() for p in produtos.split('|')]

    # Check each product
    for product in product_list:
        # Normalize: remove accents and convert to lowercase
        normalized = remove_accents(product).lower().strip()

        if normalized in EXCLUDED_PRODUCTS:
            return True

    return False


def load_slack_config(config_file: str = SLACK_CONFIG_FILE) -> dict:
    """
    Load Slack configuration from JSON file.

    Args:
        config_file: Path to the configuration file

    Returns:
        Slack configuration dictionary

    Raises:
        FileNotFoundError: If config file doesn't exist
    """
    # Use current working directory for config files (for standalone binary)
    config_path = Path.cwd() / config_file

    if not config_path.exists():
        raise FileNotFoundError(f"Slack configuration file not found: {config_path}")

    with open(config_path, 'r') as f:
        config = json.load(f)

    return config['slack']


def get_channel_id(bot_token: str, channel_name: str) -> Optional[str]:
    """
    Get channel ID from channel name.

    Args:
        bot_token: Slack bot token
        channel_name: Channel name (with or without #)

    Returns:
        Channel ID or None if not found
    """
    try:
        # Remove # if present
        channel_name_clean = channel_name.lstrip('#')

        # Try public channels first
        response = requests.get(
            'https://slack.com/api/conversations.list',
            headers={'Authorization': f'Bearer {bot_token}'},
            params={
                'types': 'public_channel',
                'limit': 1000
            },
            timeout=10
        )

        result = response.json()

        if result.get('ok'):
            for channel in result.get('channels', []):
                if channel.get('name') == channel_name_clean:
                    return channel.get('id')

        # Try private channels
        response = requests.get(
            'https://slack.com/api/conversations.list',
            headers={'Authorization': f'Bearer {bot_token}'},
            params={
                'types': 'private_channel',
                'limit': 1000
            },
            timeout=10
        )

        result = response.json()

        if result.get('ok'):
            for channel in result.get('channels', []):
                if channel.get('name') == channel_name_clean:
                    return channel.get('id')

        return None

    except Exception as e:
        print(f"  Warning: Could not get channel ID: {e}")
        return None


def get_channel_members(bot_token: str, channel_id: str) -> List[str]:
    """
    Get list of user IDs in a channel.

    Args:
        bot_token: Slack bot token
        channel_id: Channel ID

    Returns:
        List of user IDs
    """
    try:
        response = requests.get(
            'https://slack.com/api/conversations.members',
            headers={'Authorization': f'Bearer {bot_token}'},
            params={'channel': channel_id},
            timeout=10
        )

        result = response.json()

        if result.get('ok'):
            return result.get('members', [])

        return []

    except Exception as e:
        print(f"  Warning: Could not get channel members: {e}")
        return []


def get_user_info(bot_token: str, user_id: str) -> Optional[Dict[str, Any]]:
    """
    Get user information from Slack.

    Args:
        bot_token: Slack bot token
        user_id: User ID

    Returns:
        User info dictionary or None
    """
    try:
        response = requests.get(
            'https://slack.com/api/users.info',
            headers={'Authorization': f'Bearer {bot_token}'},
            params={'user': user_id},
            timeout=10
        )

        result = response.json()

        if result.get('ok'):
            return result.get('user')

        return None

    except Exception as e:
        print(f"  Warning: Could not get user info: {e}")
        return None


def match_operator_to_slack_user(operator_name: str,
                                  slack_users: List[Dict[str, Any]]) -> Optional[str]:
    """
    Match operator name to Slack user ID.

    Tries to match by:
    1. Real name (exact match, case-insensitive)
    2. Display name (exact match, case-insensitive)
    3. Real name (partial match, case-insensitive)

    Args:
        operator_name: Operator name from Excel
        slack_users: List of Slack user info dictionaries

    Returns:
        User ID if match found, None otherwise
    """
    if not operator_name:
        return None

    # Normalize operator name
    operator_normalized = remove_accents(operator_name).lower().strip()

    # Try exact matches first
    for user in slack_users:
        profile = user.get('profile', {})
        real_name = profile.get('real_name', '')
        display_name = profile.get('display_name', '')

        # Normalize Slack names
        real_name_normalized = remove_accents(real_name).lower().strip()
        display_name_normalized = remove_accents(display_name).lower().strip()

        # Exact match on real name or display name
        if (operator_normalized == real_name_normalized or
            operator_normalized == display_name_normalized):
            return user.get('id')

    # Try partial matches
    for user in slack_users:
        profile = user.get('profile', {})
        real_name = profile.get('real_name', '')

        real_name_normalized = remove_accents(real_name).lower().strip()

        # Check if operator name is in real name or vice versa
        if (operator_normalized in real_name_normalized or
            real_name_normalized in operator_normalized):
            return user.get('id')

    return None


class RepurchaseData:
    """Class to represent repurchase data for a client."""
    
    def __init__(self, row_data: Dict[str, Any]):
        """Initialize repurchase data from a row dictionary."""
        self.gerente = row_data.get('Gerente')
        self.operador = row_data.get('Operador')
        self.grupo = row_data.get('Grupo')
        self.cedente = row_data.get('Cedente')
        self.particularidade = row_data.get('Particularidade')
        self.tipo_recompra = row_data.get('Tipo de Recompra')
        self.prazo_recompra = row_data.get('Prazo de recompra')
        self.prazo_rating = row_data.get('Prazo de recompra definido pelo Rating')
        self.aumentar_prazo = row_data.get('Aumentar prazo de recompra para quantos dias?')
        self.contato_cedente = row_data.get('Contato do Cedente')
        self.restricao_operador = row_data.get('Restrição em mudar de operador? Sim ou Não')
        self.observacao_cliente = row_data.get('OBSERVAÇÃO DO CLIENTE. \nALGUMA INFORMAÇÃO QUE FUJA DO PADRÃO')
        self.parecer_resumido = row_data.get('Parecer resumido do cliente')
        self.rating = row_data.get('Rating 03/09/2024')
        self.ferias_coletivas = row_data.get('FÉRIAS COLETIVAS')
    
    def __repr__(self):
        """String representation of the repurchase data."""
        return f"RepurchaseData(cedente={self.cedente}, gerente={self.gerente}, tipo={self.tipo_recompra})"
    
    def to_dict(self) -> Dict[str, Any]:
        """Convert to dictionary."""
        return {
            'gerente': self.gerente,
            'operador': self.operador,
            'grupo': self.grupo,
            'cedente': self.cedente,
            'particularidade': self.particularidade,
            'tipo_recompra': self.tipo_recompra,
            'prazo_recompra': self.prazo_recompra,
            'prazo_rating': self.prazo_rating,
            'aumentar_prazo': self.aumentar_prazo,
            'contato_cedente': self.contato_cedente,
            'restricao_operador': self.restricao_operador,
            'observacao_cliente': self.observacao_cliente,
            'parecer_resumido': self.parecer_resumido,
            'rating': self.rating,
            'ferias_coletivas': self.ferias_coletivas,
        }


def read_excel_file(file_path: str = EXCEL_FILE_PATH, 
                    sheet_name: str = MAIN_SHEET_NAME) -> List[RepurchaseData]:
    """
    Read the Excel file and return a list of RepurchaseData objects.
    
    Args:
        file_path: Path to the Excel file
        sheet_name: Name of the sheet to read
        
    Returns:
        List of RepurchaseData objects
        
    Raises:
        FileNotFoundError: If the Excel file doesn't exist
        ValueError: If the sheet doesn't exist
    """
    # Check if file exists
    if not Path(file_path).exists():
        raise FileNotFoundError(f"Excel file not found: {file_path}")
    
    print(f"Reading Excel file: {file_path}")
    print(f"Sheet: {sheet_name}")
    print()
    
    # Load workbook
    wb = openpyxl.load_workbook(file_path, data_only=True)
    
    # Check if sheet exists
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found. Available sheets: {wb.sheetnames}")
    
    ws = wb[sheet_name]
    
    print(f"✓ Workbook loaded successfully")
    print(f"  Total rows: {ws.max_row}")
    print(f"  Total columns: {ws.max_column}")
    print()
    
    # Read headers from first row
    headers = []
    for col in range(1, ws.max_column + 1):
        header = ws.cell(1, col).value
        headers.append(header)
    
    print(f"Headers found: {len(headers)}")
    for i, header in enumerate(headers, 1):
        if header:
            print(f"  Col {i}: {header}")
    print()
    
    # Read data rows
    repurchase_data_list = []
    for row_num in range(2, ws.max_row + 1):
        row_data = {}
        for col_num, header in enumerate(headers, 1):
            if header:
                cell_value = ws.cell(row_num, col_num).value
                row_data[header] = cell_value
        
        # Only add rows that have at least a cedente (client name)
        if row_data.get('Cedente'):
            repurchase_data = RepurchaseData(row_data)
            repurchase_data_list.append(repurchase_data)
    
    print(f"✓ Data loaded successfully")
    print(f"  Total records: {len(repurchase_data_list)}")
    print()

    return repurchase_data_list


def display_summary(data_list: List[RepurchaseData]) -> None:
    """
    Display a summary of the repurchase data.

    Args:
        data_list: List of RepurchaseData objects
    """
    print("=" * 100)
    print("REPURCHASE DATA SUMMARY")
    print("=" * 100)
    print()

    # Count by manager
    managers = {}
    for data in data_list:
        manager = data.gerente or "N/A"
        managers[manager] = managers.get(manager, 0) + 1

    print("Records by Manager:")
    for manager, count in sorted(managers.items(), key=lambda x: x[1], reverse=True):
        print(f"  {manager}: {count}")
    print()

    # Count by repurchase type
    repurchase_types = {}
    for data in data_list:
        rep_type = data.tipo_recompra or "N/A"
        repurchase_types[rep_type] = repurchase_types.get(rep_type, 0) + 1

    print("Records by Repurchase Type:")
    for rep_type, count in sorted(repurchase_types.items(), key=lambda x: x[1], reverse=True):
        print(f"  {rep_type}: {count}")
    print()

    # Count by operator
    operators = {}
    for data in data_list:
        operator = data.operador or "N/A"
        operators[operator] = operators.get(operator, 0) + 1

    print("Records by Operator:")
    for operator, count in sorted(operators.items(), key=lambda x: x[1], reverse=True):
        print(f"  {operator}: {count}")
    print()

    print("=" * 100)


def display_sample_records(data_list: List[RepurchaseData], num_records: int = 5) -> None:
    """
    Display sample records from the data.

    Args:
        data_list: List of RepurchaseData objects
        num_records: Number of records to display
    """
    print("=" * 100)
    print(f"SAMPLE RECORDS (First {num_records})")
    print("=" * 100)
    print()

    for i, data in enumerate(data_list[:num_records], 1):
        print(f"Record {i}:")
        print(f"  Cedente: {data.cedente}")
        print(f"  Gerente: {data.gerente}")
        print(f"  Operador: {data.operador}")
        print(f"  Grupo: {data.grupo}")
        print(f"  Tipo de Recompra: {data.tipo_recompra}")
        print(f"  Prazo de recompra: {data.prazo_recompra}")
        print(f"  Prazo Rating: {data.prazo_rating}")
        if data.particularidade:
            particularidade_preview = str(data.particularidade)[:100]
            if len(str(data.particularidade)) > 100:
                particularidade_preview += "..."
            print(f"  Particularidade: {particularidade_preview}")
        print()

    print("=" * 100)


def remove_accents(text: str) -> str:
    """
    Remove accents from a string.

    Args:
        text: Input string

    Returns:
        String without accents
    """
    if not text:
        return ""

    # Normalize to NFD (decomposed form) and filter out combining characters
    nfd = unicodedata.normalize('NFD', text)
    return ''.join(char for char in nfd if unicodedata.category(char) != 'Mn')


def filter_by_repurchase_type(data_list: List[RepurchaseData],
                               types: List[str] = None) -> List[RepurchaseData]:
    """
    Filter records by repurchase type (Tipo de Recompra).
    Matches with or without accents.

    Args:
        data_list: List of RepurchaseData objects
        types: List of repurchase types to filter (default: ["Operacao", "Bordero"])

    Returns:
        List of matching RepurchaseData objects
    """
    if types is None:
        types = ["Operacao", "Bordero"]

    # Normalize the search types (remove accents and convert to lowercase)
    normalized_types = [remove_accents(t).lower() for t in types]

    matches = []
    for data in data_list:
        if data.tipo_recompra:
            # Normalize the data value
            normalized_value = remove_accents(str(data.tipo_recompra)).lower().strip()

            # Check if any of the search types match
            for search_type in normalized_types:
                if search_type in normalized_value:
                    matches.append(data)
                    break  # Only add once even if multiple types match

    return matches


def query_apr_capa_for_date(target_date: date = None) -> List[Dict[str, Any]]:
    """
    Query APR_CAPA joined with cedente table for records created after 16h15 of previous BUSINESS day.
    Also includes product information from APR_TITULOS -> ProdutoCedente -> ProdutoAtributo -> Produto.

    New logic with business days:
    - If today is a business day: query from 16:15 of previous business day
    - If today is NOT a business day (weekend/holiday): skip query (no alerts)
    - Only include records with tipo_operacao = 1

    Example scenarios:
    - Monday: Query from Friday 16:15 onwards
    - Tuesday-Friday: Query from previous day 16:15 onwards
    - Saturday/Sunday: Don't query (not a business day)
    - Holiday: Don't query (not a business day)

    Args:
        target_date: Date to query (default: today)

    Returns:
        List of dictionaries with APR_CAPA, cedente, and product data, or empty list if not a business day
    """
    if target_date is None:
        target_date = date.today()

    # Check if today is a business day
    if not is_business_day(target_date):
        print(f"⚠ {target_date} is NOT a business day (weekend or holiday)")
        print(f"⚠ Skipping query - no alerts will be sent")
        print()
        return []

    # Get previous business day
    previous_business_day = get_previous_business_day(target_date)

    if previous_business_day is None:
        print(f"⚠ Could not find previous business day from {target_date}")
        print(f"⚠ Using previous calendar day as fallback")
        previous_business_day = target_date - timedelta(days=1)

    # Calculate the cutoff datetime: previous business day at 16:15:00
    cutoff_datetime = datetime.combine(previous_business_day, time(16, 15, 0))

    print(f"Today: {target_date} (BUSINESS DAY)")
    print(f"Previous business day: {previous_business_day}")
    print(f"Querying APR_CAPA for records created after: {cutoff_datetime}")
    print(f"Filter: tipo_operacao = 1")
    print()

    # Create database session
    engine = create_db_engine()
    Session = sessionmaker(bind=engine)
    session = Session()

    try:
        # First, query APR_CAPA joined with cedente
        # APR_CAPA.CEDENTE corresponds to cedente.apelido
        # Filter: dataentrada >= cutoff_datetime AND tipo_operacao = 1
        results = session.query(
            APRCapa.DATA,
            APRCapa.NUMERO,
            APRCapa.CEDENTE,
            APRCapa.GERENTE,
            APRCapa.BORDERO,
            APRCapa.QTD_PROPOSTOS,
            APRCapa.VLR_PROPOSTOS,
            APRCapa.QTD_APROVADOS,
            APRCapa.VLR_APROVADOS,
            APRCapa.empresa,
            APRCapa.tipo_operacao,
            APRCapa.dataentrada,
            Cedente.nome.label('cedente_nome'),
            Cedente.grupo.label('cedente_grupo'),
            Cedente.gerente.label('cedente_gerente')
        ).outerjoin(
            Cedente,
            APRCapa.CEDENTE == Cedente.apelido
        ).filter(
            APRCapa.dataentrada >= cutoff_datetime,
            APRCapa.tipo_operacao == 1
        ).all()

        print(f"✓ Found {len(results)} records in APR_CAPA for {target_date}")
        print()

        # Now query products for each borderô using raw SQL (more efficient for aggregation)
        # We need to use pymssql directly for this complex query
        import pymssql
        config_path = Path.cwd() / DATABASE_CONFIG_FILE
        with open(config_path, 'r') as f:
            config = json.load(f)

        db_config = config['databases']['mssql']

        conn = pymssql.connect(
            server=db_config['server'],
            port=db_config['port'],
            user=db_config['user'],
            password=db_config['password'],
            database=db_config['scheme']
        )

        cursor = conn.cursor(as_dict=True)

        # Query to get products per borderô
        # Use STUFF + FOR XML PATH to concatenate distinct product names
        cursor.execute('''
            SELECT
                c.NUMERO,
                STUFF((
                    SELECT DISTINCT ' | ' + p.Descritivo
                    FROM APR_TITULOS t2 WITH (NOLOCK)
                    LEFT JOIN ProdutoCedente pc2 WITH (NOLOCK)
                        ON t2.id_produto = pc2.Id
                    LEFT JOIN ProdutoAtributo pa2 WITH (NOLOCK)
                        ON pc2.IdProdutoAtributo = pa2.Id
                    LEFT JOIN Produto p WITH (NOLOCK)
                        ON pa2.IdProduto = p.Id
                    WHERE t2.NUMERO = c.NUMERO
                      AND t2.DATA = c.DATA
                      AND p.Descritivo IS NOT NULL
                    FOR XML PATH(''), TYPE
                ).value('.', 'NVARCHAR(MAX)'), 1, 3, '') AS produtos
            FROM APR_capa c WITH (NOLOCK)
            WHERE c.dataentrada >= %s
              AND c.tipo_operacao = 1
        ''', (cutoff_datetime,))

        produtos_por_bordero = {}
        for row in cursor.fetchall():
            if row['produtos']:
                produtos_por_bordero[row['NUMERO']] = row['produtos']

        cursor.close()
        conn.close()

        # Convert to list of dictionaries
        result_list = []
        for row in results:
            numero = row.NUMERO
            produtos = produtos_por_bordero.get(numero, None)

            result_list.append({
                'DATA': row.DATA,
                'NUMERO': row.NUMERO,
                'CEDENTE': row.CEDENTE,  # This is the apelido
                'GERENTE': row.GERENTE,
                'BORDERO': row.BORDERO,
                'QTD_PROPOSTOS': row.QTD_PROPOSTOS,
                'VLR_PROPOSTOS': row.VLR_PROPOSTOS,
                'QTD_APROVADOS': row.QTD_APROVADOS,
                'VLR_APROVADOS': row.VLR_APROVADOS,
                'empresa': row.empresa,
                'tipo_operacao': row.tipo_operacao,
                'dataentrada': row.dataentrada,
                'cedente_nome': row.cedente_nome,
                'cedente_grupo': row.cedente_grupo,
                'cedente_gerente': row.cedente_gerente,
                'produtos': produtos,  # NEW: Product names (pipe-separated)
            })

        return result_list

    finally:
        session.close()


def cross_reference_data(excel_data: List[RepurchaseData],
                         apr_data: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """
    Cross-reference Excel repurchase data with APR_CAPA data.

    Matches Excel column "Cedente" with APR_CAPA column "CEDENTE" (which is cedente.apelido).

    Args:
        excel_data: List of RepurchaseData objects from Excel
        apr_data: List of dictionaries from APR_CAPA query

    Returns:
        List of dictionaries with combined data
    """
    print("Cross-referencing Excel data with APR_CAPA data...")
    print()

    # Create a lookup dictionary from Excel data
    # Key: cedente name (normalized), Value: RepurchaseData object
    excel_lookup = {}
    for data in excel_data:
        if data.cedente:
            # Normalize the cedente name (remove accents, lowercase, strip)
            normalized_name = remove_accents(data.cedente).lower().strip()
            excel_lookup[normalized_name] = data

    # Cross-reference
    matched_records = []
    unmatched_apr = []

    for apr_record in apr_data:
        cedente_apelido = apr_record['CEDENTE']

        if cedente_apelido:
            # Normalize the apelido
            normalized_apelido = remove_accents(cedente_apelido).lower().strip()

            # Look up in Excel data
            excel_record = excel_lookup.get(normalized_apelido)

            if excel_record:
                # Match found - combine the data
                combined = {
                    # APR_CAPA data
                    'apr_data': apr_record['DATA'],
                    'apr_numero': apr_record['NUMERO'],
                    'apr_cedente': apr_record['CEDENTE'],
                    'apr_gerente': apr_record['GERENTE'],
                    'apr_bordero': apr_record['BORDERO'],
                    'apr_qtd_propostos': apr_record['QTD_PROPOSTOS'],
                    'apr_vlr_propostos': apr_record['VLR_PROPOSTOS'],
                    'apr_qtd_aprovados': apr_record['QTD_APROVADOS'],
                    'apr_vlr_aprovados': apr_record['VLR_APROVADOS'],
                    'apr_empresa': apr_record['empresa'],
                    'apr_tipo_operacao': apr_record['tipo_operacao'],
                    'apr_dataentrada': apr_record['dataentrada'],
                    'apr_produtos': apr_record.get('produtos'),  # NEW: Product names

                    # Excel repurchase data
                    'excel_cedente': excel_record.cedente,
                    'excel_gerente': excel_record.gerente,
                    'excel_operador': excel_record.operador,
                    'excel_grupo': excel_record.grupo,
                    'excel_tipo_recompra': excel_record.tipo_recompra,
                    'excel_prazo_recompra': excel_record.prazo_recompra,
                    'excel_prazo_rating': excel_record.prazo_rating,
                    'excel_particularidade': excel_record.particularidade,
                    'excel_contato_cedente': excel_record.contato_cedente,
                    'excel_restricao_operador': excel_record.restricao_operador,
                    'excel_observacao_cliente': excel_record.observacao_cliente,
                }
                matched_records.append(combined)
            else:
                # No match in Excel
                unmatched_apr.append(apr_record)

    print(f"✓ Cross-reference complete:")
    print(f"  Matched records: {len(matched_records)}")
    print(f"  Unmatched APR_CAPA records: {len(unmatched_apr)}")
    print()

    return matched_records


def load_sent_alerts() -> Dict[str, List[str]]:
    """
    Load the sent alerts tracking file.

    Returns:
        Dictionary with date as key and list of alert IDs as value
        Format: {"2026-01-19": ["METALURGICA_REUTER_2", "BRAUNAS_3", ...]}
    """
    # Use current working directory for data files (for standalone binary)
    sent_alerts_path = Path.cwd() / SENT_ALERTS_FILE

    if not sent_alerts_path.exists():
        return {}

    try:
        with open(sent_alerts_path, 'r', encoding='utf-8') as f:
            return json.load(f)
    except Exception as e:
        print(f"⚠ Warning: Could not load sent alerts file: {e}")
        return {}


def save_sent_alerts(sent_alerts: Dict[str, List[str]]) -> None:
    """
    Save the sent alerts tracking file.

    Args:
        sent_alerts: Dictionary with date as key and list of alert IDs as value
    """
    # Use current working directory for data files (for standalone binary)
    sent_alerts_path = Path.cwd() / SENT_ALERTS_FILE

    try:
        with open(sent_alerts_path, 'w', encoding='utf-8') as f:
            json.dump(sent_alerts, f, indent=2, ensure_ascii=False)
    except Exception as e:
        print(f"⚠ Warning: Could not save sent alerts file: {e}")


def cleanup_old_alerts(sent_alerts: Dict[str, List[str]], days_to_keep: int = 7) -> Dict[str, List[str]]:
    """
    Remove alerts older than specified days to keep the file clean.

    Args:
        sent_alerts: Dictionary with date as key and list of alert IDs as value
        days_to_keep: Number of days to keep in history (default: 7)

    Returns:
        Cleaned dictionary with only recent alerts
    """
    today = date.today()
    cleaned = {}

    for date_str, alerts in sent_alerts.items():
        try:
            alert_date = datetime.strptime(date_str, '%Y-%m-%d').date()
            days_diff = (today - alert_date).days

            if days_diff <= days_to_keep:
                cleaned[date_str] = alerts
        except ValueError:
            # Skip invalid date entries
            continue

    return cleaned


def get_alert_time_window() -> str:
    """
    Get the current alert time window based on current time.

    Windows:
    - "morning": 08:00 - 11:59 (send once)
    - "afternoon": 12:00 - 16:15 (send once)
    - "after_hours": 16:15+ (don't send, wait for next day)

    Returns:
        Time window identifier: "morning", "afternoon", or "after_hours"
    """
    now = datetime.now().time()

    if time(8, 0) <= now < time(12, 0):
        return "morning"
    elif time(12, 0) <= now < time(16, 15):
        return "afternoon"
    else:
        return "after_hours"


def create_alert_id(cedente: str, numero: int = None, time_window: str = None) -> str:
    """
    Create a unique alert ID for duplicity control with time window.

    New format includes time window to allow one alert before 12h and one after 12h.

    IMPORTANT: Alert ID is based on CEDENTE + TIME_WINDOW only (not borderô number).
    This ensures maximum 2 alerts per cedente per day (one morning, one afternoon),
    regardless of how many borderôs the cedente has.

    Args:
        cedente: Client name (apelido)
        numero: Borderô number (DEPRECATED - kept for backward compatibility, not used)
        time_window: Time window ("morning" or "afternoon"). If None, uses current time.

    Returns:
        Unique alert ID (e.g., "METALURGICA_REUTER_morning")
    """
    if time_window is None:
        time_window = get_alert_time_window()

    # Normalize cedente name: remove spaces, convert to uppercase
    normalized_cedente = cedente.replace(' ', '_').upper()

    # Alert ID is based on CEDENTE + TIME_WINDOW only (not borderô number)
    # This ensures max 2 alerts per cedente per day
    return f"{normalized_cedente}_{time_window}"


def is_alert_already_sent(alert_id: str, target_date: date, sent_alerts: Dict[str, List[str]]) -> bool:
    """
    Check if an alert was already sent for this time window today.

    Args:
        alert_id: Unique alert ID (includes time window)
        target_date: Date to check
        sent_alerts: Dictionary of sent alerts

    Returns:
        True if alert was already sent in this time window today, False otherwise
    """
    date_str = target_date.strftime('%Y-%m-%d')
    return date_str in sent_alerts and alert_id in sent_alerts[date_str]


def mark_alert_as_sent(alert_id: str, target_date: date, sent_alerts: Dict[str, List[str]]) -> Dict[str, List[str]]:
    """
    Mark an alert as sent for the given date and time window.

    Args:
        alert_id: Unique alert ID (includes time window)
        target_date: Date when alert was sent
        sent_alerts: Dictionary of sent alerts

    Returns:
        Updated sent_alerts dictionary
    """
    date_str = target_date.strftime('%Y-%m-%d')

    if date_str not in sent_alerts:
        sent_alerts[date_str] = []

    if alert_id not in sent_alerts[date_str]:
        sent_alerts[date_str].append(alert_id)

    return sent_alerts


def is_within_alert_hours() -> bool:
    """
    Check if current time is within alert sending hours (08:00 - 16:15).

    Returns:
        True if within alert hours, False otherwise
    """
    now = datetime.now().time()
    return time(8, 0) <= now < time(16, 15)


def send_repurchase_alerts(matched_records: List[Dict[str, Any]],
                           dry_run: bool = True) -> Dict[str, Any]:
    """
    Send Slack alerts for repurchase records.

    For each matched record, sends a message to the operacionalxcobranca channel
    mentioning the operator.

    New rules:
    - Only send alerts on BUSINESS DAYS (not weekends or holidays)
    - Only send alerts between 08:00 and 16:15
    - Send once in morning window (08:00-11:59)
    - Send once in afternoon window (12:00-16:15)
    - Records created after 16:15 will be sent next business day

    Message format: @{operator} {cedente_apelido}

    Args:
        matched_records: List of matched records from cross-reference
        dry_run: If True, only simulate sending (don't actually send)

    Returns:
        Dictionary with statistics about sent alerts
    """
    print("=" * 120)
    print(f"SENDING REPURCHASE ALERTS {'(DRY RUN)' if dry_run else ''}")
    print("=" * 120)
    print()

    # Check if today is a business day
    today = date.today()
    if not is_business_day(today):
        print(f"⚠ Today ({today}) is NOT a business day (weekend or holiday)")
        print(f"⚠ Alerts will not be sent. Run this script on a business day.")
        print()
        return {
            'total_records': len(matched_records),
            'alerts_sent': 0,
            'alerts_failed': 0,
            'alerts_skipped_duplicate': 0,
            'alerts_skipped_outside_hours': 0,
            'alerts_skipped_not_business_day': len(matched_records),
            'operator_matched': 0,
            'operator_not_matched': 0
        }

    # Check if we're within alert hours
    if not is_within_alert_hours():
        current_time = datetime.now().strftime('%H:%M:%S')
        print(f"⚠ Current time ({current_time}) is outside alert hours (08:00 - 16:15)")
        print(f"⚠ Alerts will not be sent. Run this script between 08:00 and 16:15.")
        print()
        return {
            'total_records': len(matched_records),
            'alerts_sent': 0,
            'alerts_failed': 0,
            'alerts_skipped_duplicate': 0,
            'alerts_skipped_outside_hours': len(matched_records),
            'alerts_skipped_not_business_day': 0,
            'operator_matched': 0,
            'operator_not_matched': 0
        }

    # Get current time window
    time_window = get_alert_time_window()
    print(f"Today is a BUSINESS DAY: {today}")
    print(f"Current time window: {time_window}")
    print()

    stats = {
        'total_records': len(matched_records),
        'alerts_sent': 0,
        'alerts_failed': 0,
        'alerts_skipped_duplicate': 0,
        'alerts_skipped_outside_hours': 0,
        'alerts_skipped_excluded_product': 0,
        'operator_matched': 0,
        'operator_not_matched': 0,
    }

    if not matched_records:
        print("No matched records to send alerts for.")
        return stats

    # Load sent alerts tracking
    sent_alerts = load_sent_alerts()
    today = date.today()

    # Clean up old alerts (keep last 7 days)
    sent_alerts = cleanup_old_alerts(sent_alerts, days_to_keep=7)

    print(f"Duplicity Control:")
    date_str = today.strftime('%Y-%m-%d')
    alerts_today = sent_alerts.get(date_str, [])
    print(f"  Alerts already sent today ({date_str}): {len(alerts_today)}")
    print()

    try:
        # Load Slack config
        slack_config = load_slack_config()
        bot_token = slack_config.get('bot_token')

        if not bot_token:
            print("❌ Slack bot token not configured")
            return stats

        # Get channel ID
        print(f"Looking up channel: #{SLACK_CHANNEL_NAME}")
        channel_id = get_channel_id(bot_token, SLACK_CHANNEL_NAME)

        if not channel_id:
            print(f"❌ Channel '#{SLACK_CHANNEL_NAME}' not found")
            return stats

        print(f"✓ Channel ID: {channel_id}")
        print()

        # Get channel members
        print("Fetching channel members...")
        member_ids = get_channel_members(bot_token, channel_id)
        print(f"✓ Found {len(member_ids)} members in channel")
        print()

        # Get user info for all members
        print("Fetching user information...")
        slack_users = []
        for user_id in member_ids:
            user_info = get_user_info(bot_token, user_id)
            if user_info and not user_info.get('is_bot'):  # Exclude bots
                slack_users.append(user_info)

        print(f"✓ Retrieved info for {len(slack_users)} users (excluding bots)")
        print()

        # Display user mapping
        print("Slack users in channel:")
        for user in slack_users[:20]:  # Show first 20 users
            profile = user.get('profile', {})
            print(f"  - {profile.get('real_name', 'N/A')} (@{user.get('name', 'N/A')}) - ID: {user.get('id')}")
        if len(slack_users) > 20:
            print(f"  ... and {len(slack_users) - 20} more users")
        print()

        print(f"Target channel: #{SLACK_CHANNEL_NAME} (ID: {channel_id})")
        print()

        # Process each matched record
        print("=" * 120)
        print("PROCESSING ALERTS")
        print("=" * 120)
        print()

        for i, record in enumerate(matched_records, 1):
            operator_name = record.get('excel_operador')
            cedente_apelido = record.get('apr_cedente')
            numero = record.get('apr_numero')
            tipo_operacao = record.get('apr_tipo_operacao')
            dataentrada = record.get('apr_dataentrada')
            produtos = record.get('apr_produtos')  # Get product information
            empresa = record.get('apr_empresa')  # Get company information

            print(f"{i}. Cedente: {cedente_apelido} | Borderô: {numero} | Operador: {operator_name} | tipo_operacao: {tipo_operacao} | Empresa: {empresa}")
            if dataentrada:
                print(f"   dataentrada: {dataentrada.strftime('%Y-%m-%d %H:%M:%S')}")
            if produtos:
                print(f"   Produtos: {produtos}")

            # Check if product should be excluded from alerts
            if should_exclude_product(produtos):
                stats['alerts_skipped_excluded_product'] = stats.get('alerts_skipped_excluded_product', 0) + 1
                print(f"   ⏭ Product excluded from alerts - SKIPPED")
                print()
                continue

            # Create unique alert ID for duplicity control (includes time window)
            alert_id = create_alert_id(cedente_apelido, numero, time_window)

            # Check if alert was already sent in this time window today
            if is_alert_already_sent(alert_id, today, sent_alerts):
                stats['alerts_skipped_duplicate'] += 1
                print(f"   ⏭ Alert already sent in {time_window} window - SKIPPED")
                print()
                continue

            # Match operator to Slack user
            user_id = match_operator_to_slack_user(operator_name, slack_users)

            if user_id:
                stats['operator_matched'] += 1

                # Format message with user mention
                # Slack format for mentioning: <@USER_ID>
                # Message format: @{operator} {cedente}
                # For BMA company, append "BMA Sec" at the end
                # NOTE: Products are NOT included in the message
                message = f"<@{user_id}> {cedente_apelido}"

                # Append "BMA Sec" for BMA company records
                if empresa == "BMA":
                    message += " BMA Sec"

                print(f"   ✓ Matched operator to Slack user: {user_id}")
                print(f"   Message: {message}")

                if not dry_run:
                    # Send actual message
                    try:
                        response = requests.post(
                            'https://slack.com/api/chat.postMessage',
                            headers={
                                'Authorization': f'Bearer {bot_token}',
                                'Content-Type': 'application/json'
                            },
                            json={
                                'channel': channel_id,
                                'text': message
                            },
                            timeout=10
                        )

                        result = response.json()

                        if result.get('ok'):
                            print(f"   ✓ Alert sent successfully")
                            stats['alerts_sent'] += 1
                            # Mark alert as sent
                            sent_alerts = mark_alert_as_sent(alert_id, today, sent_alerts)
                        else:
                            print(f"   ✗ Failed to send alert: {result.get('error')}")
                            stats['alerts_failed'] += 1

                    except Exception as e:
                        print(f"   ✗ Error sending alert: {e}")
                        stats['alerts_failed'] += 1
                else:
                    print(f"   [DRY RUN] Would send alert")
                    stats['alerts_sent'] += 1
                    # Mark alert as sent even in dry-run mode for duplicity control testing
                    sent_alerts = mark_alert_as_sent(alert_id, today, sent_alerts)
            else:
                stats['operator_not_matched'] += 1
                print(f"   ⚠ Could not match operator '{operator_name}' to Slack user")

            print()

        # Save sent alerts to file
        save_sent_alerts(sent_alerts)

        # Print summary
        print("=" * 120)
        print("ALERT SUMMARY")
        print("=" * 120)
        print(f"Time window: {time_window}")
        print(f"Total records: {stats['total_records']}")
        print(f"Operators matched: {stats['operator_matched']}")
        print(f"Operators not matched: {stats['operator_not_matched']}")
        print(f"Alerts sent: {stats['alerts_sent']}")
        print(f"Alerts failed: {stats['alerts_failed']}")
        print(f"Alerts skipped (duplicate in this window): {stats['alerts_skipped_duplicate']}")
        print(f"Alerts skipped (excluded product): {stats['alerts_skipped_excluded_product']}")
        print(f"Alerts skipped (outside hours): {stats['alerts_skipped_outside_hours']}")
        print()

        return stats

    except Exception as e:
        print(f"❌ Error in send_repurchase_alerts: {e}")
        import traceback
        traceback.print_exc()
        return stats


def search_by_cedente(data_list: List[RepurchaseData], cedente_name: str) -> List[RepurchaseData]:
    """
    Search for records by cedente (client) name.

    Args:
        data_list: List of RepurchaseData objects
        cedente_name: Name of the cedente to search for (case-insensitive, partial match)

    Returns:
        List of matching RepurchaseData objects
    """
    cedente_name_lower = cedente_name.lower()
    matches = []

    for data in data_list:
        if data.cedente and cedente_name_lower in data.cedente.lower():
            matches.append(data)

    return matches


def main():
    """Main function to demonstrate the Excel reading and database cross-referencing."""
    print("=" * 120)
    print("REPURCHASE REPORT - Excel + Database Cross-Reference")
    print("=" * 120)
    print()

    try:
        # Read the Excel file
        repurchase_data = read_excel_file()

        # Display summary
        display_summary(repurchase_data)
        print()

        # Filter by repurchase type: "Operacao" or "Bordero"
        filtered_data = filter_by_repurchase_type(repurchase_data, ["Operacao", "Bordero"])

        print("=" * 120)
        print(f"FILTERED RECORDS - Tipo de Recompra: 'Operacao' or 'Bordero'")
        print("=" * 120)
        print(f"Total filtered records: {len(filtered_data)}")
        print()

        # Show breakdown of filtered records
        if filtered_data:
            tipo_counts = {}
            for data in filtered_data:
                tipo = data.tipo_recompra or "N/A"
                tipo_counts[tipo] = tipo_counts.get(tipo, 0) + 1

            print("Breakdown by Tipo de Recompra:")
            for tipo, count in sorted(tipo_counts.items(), key=lambda x: x[1], reverse=True):
                print(f"  {tipo}: {count}")
            print()

        print("=" * 120)
        print()

        # Query APR_CAPA for current date
        print("=" * 120)
        print("QUERYING APR_CAPA DATABASE")
        print("=" * 120)
        print()

        apr_data = query_apr_capa_for_date()

        if apr_data:
            print(f"Sample APR_CAPA records (first 5):")
            print("-" * 120)
            for i, record in enumerate(apr_data[:5], 1):
                print(f"{i}. CEDENTE: {record['CEDENTE']}")
                print(f"   NUMERO: {record['NUMERO']} | GERENTE: {record['GERENTE']}")
                print(f"   QTD_PROPOSTOS: {record['QTD_PROPOSTOS']} | VLR_PROPOSTOS: R$ {record['VLR_PROPOSTOS']:,.2f}")
                print(f"   QTD_APROVADOS: {record['QTD_APROVADOS']} | VLR_APROVADOS: R$ {record['VLR_APROVADOS']:,.2f}")

                # Display tipo_operacao
                tipo_op = record.get('tipo_operacao')
                tipo_op_str = str(tipo_op) if tipo_op is not None else 'N/A'
                print(f"   TIPO_OPERACAO: {tipo_op_str}", end='')

                # Display dataentrada
                dataentrada = record.get('dataentrada')
                if dataentrada:
                    dataentrada_str = dataentrada.strftime('%Y-%m-%d %H:%M:%S')
                    print(f" | DATAENTRADA: {dataentrada_str}")
                else:
                    print(f" | DATAENTRADA: N/A")

                print()

        print("=" * 120)
        print()

        # Cross-reference Excel data with APR_CAPA data
        # Only use filtered Excel data (Operacao/Bordero)
        print("=" * 120)
        print("CROSS-REFERENCING EXCEL WITH APR_CAPA")
        print("=" * 120)
        print()

        matched_records = cross_reference_data(filtered_data, apr_data)

        if matched_records:
            print(f"Matched records (first 10):")
            print("-" * 120)
            for i, record in enumerate(matched_records[:10], 1):
                print(f"{i}. CEDENTE: {record['apr_cedente']}")
                print(f"   APR: Proposta {record['apr_numero']} | Gerente: {record['apr_gerente']}")
                print(f"   APR: Aprovados: {record['apr_qtd_aprovados']} | Valor: R$ {record['apr_vlr_aprovados']:,.2f}")

                # Display tipo_operacao and dataentrada
                tipo_op = record.get('apr_tipo_operacao')
                tipo_op_str = str(tipo_op) if tipo_op is not None else 'N/A'
                dataentrada = record.get('apr_dataentrada')
                dataentrada_str = dataentrada.strftime('%Y-%m-%d %H:%M:%S') if dataentrada else 'N/A'
                print(f"   APR: Tipo Operação: {tipo_op_str} | Data Entrada: {dataentrada_str}")

                print(f"   EXCEL: Tipo Recompra: {record['excel_tipo_recompra']} | Prazo: {record['excel_prazo_recompra']}")
                print(f"   EXCEL: Operador: {record['excel_operador']}")
                if record['excel_particularidade']:
                    particularidade = str(record['excel_particularidade'])[:80]
                    if len(str(record['excel_particularidade'])) > 80:
                        particularidade += "..."
                    print(f"   EXCEL: Particularidade: {particularidade}")
                print()

        print("=" * 120)
        print()

        # Send Slack alerts (SEND by default)
        # To test without sending, pass --dry-run flag
        dry_run = '--dry-run' in sys.argv

        if matched_records:
            alert_stats = send_repurchase_alerts(matched_records, dry_run=dry_run)

        print("=" * 120)
        print()

        # Example: Search for a specific cedente
        if len(sys.argv) > 1 and sys.argv[1] not in ['--dry-run']:
            search_term = sys.argv[1]
            print(f"Searching for cedente containing: '{search_term}'")
            print()
            matches = search_by_cedente(repurchase_data, search_term)

            if matches:
                print(f"Found {len(matches)} matching record(s):")
                print()
                for i, data in enumerate(matches, 1):
                    print(f"Match {i}:")
                    print(f"  Cedente: {data.cedente}")
                    print(f"  Gerente: {data.gerente}")
                    print(f"  Operador: {data.operador}")
                    print(f"  Tipo de Recompra: {data.tipo_recompra}")
                    print(f"  Prazo de recompra: {data.prazo_recompra}")
                    if data.particularidade:
                        print(f"  Particularidade: {data.particularidade}")
                    print()
            else:
                print(f"No records found matching '{search_term}'")

        print("✓ Program completed successfully")
        return 0

    except FileNotFoundError as e:
        print(f"✗ Error: {e}", file=sys.stderr)
        return 1
    except Exception as e:
        print(f"✗ Unexpected error: {e}", file=sys.stderr)
        import traceback
        traceback.print_exc()
        return 1


if __name__ == "__main__":
    sys.exit(main())

